import datetime
import json
import os
import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DataProcessorConfig:
    """配置参数类"""
    # 定义纵向速度区间顺序（车辆）
    VEHICLE_SPEED_ORDER = ['0-30', '30-60', '60-90', '90-120', '120-150', '150-180', '0-200']
    
    # 定义纵向速度区间顺序（行人）
    PEDESTRIAN_SPEED_ORDER = ['0-30', '30-60', '60-90', '90-120', '120-150', '150-180', '0-200']

    # 定义车辆类编号集合
    VEHICLE_CLASS_SET = {0, 1, 3, 6}
    # 定义行人/骑行人编号集合
    PEDESTRIAN_CLASS_SET = {2, 4}
    
    # 车辆类字符串标识
    VEHICLE_CLASS_STR = ['小车', '大车', '其他', '车辆', 'other']
    # 行人类字符串标识
    PEDESTRIAN_CLASS_STR = ['行人', '步行人', '二轮车', '骑行人']
    
    # BugScan信号类型
    BUGSCAN_KEYS = ["纵向测距", "横向测距", "纵向测速", "横向测速"]
    
    UNIFIED_CONFIG = {
        'classifications': {
            "ALL": ['ALL'],
            "小车": [1],
            "大车": [0, 3],
            "other": [6],
            "车辆": [0, 1, 3, 6],
            "骑行人": [4],
            "步行人": [2],
        },
        # 不同类型的bins和范围配置
        'speed_config': {
            'end_limit': 200,
            'bins': [0, 30, 60, 90, 120],
            'classes': [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]  # 所有相关的class
        },
        'vehicle_config': {
            'end_limit': 200,
            'bins': [0, 30, 60, 90, 120],
            'classes': [0, 1, 3, 6]  # 车辆相关的class
        },
        'pedestrian_config': {
            'end_limit': 200,
            'bins': [0, 30, 60, 90, 120],
            'classes': [2, 4]  # 行人相关的class
        },
        'all_config': {
            'end_limit': 200,
            'bins': [0, 30, 60, 90, 120, 150, 180],
            'classes': [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]  # 所有相关的class
        }
    }

    # KPI配置（基于速度区间）
    KPI_CONFIGS = {
        'vehicle': {
            '纵向距离': {'0-30': 5.0, '30-60': 5.0, '60-90': 7.0, '90-120': 9.0, '0-200': 12.0},
            '横向距离': {'0-30': 0.3, '30-60': 0.3, '60-90': 0.6, '90-120': 1.0, '0-200': 3.0},
            '纵向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'},
            '横向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'}
        },
        'pedestrian': {
            '纵向距离': {'0-30': 5.0, '30-60': 5.0, '60-90': 7.0, '90-120': 9.0, '0-200': 12.0},
            '横向距离': {'0-30': 0.3, '30-60': 0.3, '60-90': 0.6, '90-120': 1.0, '0-200': 3.0},
            '纵向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'},
            '横向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'}
        },
        'all': {
            '纵向距离': {'0-30': 5.0, '30-60': 5.0, '60-90': 7.0, '90-120': 9.0, '0-200': 12.0},
            '横向距离': {'0-30': 0.3, '30-60': 0.3, '60-90': 0.6, '90-120': 1.0, '0-200': 3.0},
            '纵向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'},
            '横向速度': {'0-30': 'max(10%, 1)', '30-60': 'max(10%, 1)', '60-90': 'max(10%, 1)', '90-120': 'max(10%, 1)', '0-200': 'max(10%, 1)'}
        }
    }
    
    # 字段映射
    FIELD_TO_KPI_KEY = {
        'x_offset': '纵向距离',
        'y_offset': '横向距离',
        'vx_offset': '纵向速度',
        'vy_offset': '横向速度'
    }
    
    FIELD_TO_STATUS_COL = {
        'x_offset': '纵向距离_KPI达标状态',
        'y_offset': '横向距离_KPI达标状态',
        'vx_offset': '纵向速度_KPI达标状态',
        'vy_offset': '横向速度_KPI达标状态'
    }
    
    FIELD_TO_KPI_COL = {
        'x_offset': '纵向距离_KPI需求值',
        'y_offset': '横向距离_KPI需求值',
        'vx_offset': '纵向速度_KPI需求值',
        'vy_offset': '横向速度_KPI需求值'
    }
    
    # class到KPI类型映射
    CLASS_TO_KPI_TYPE = {
        0: 'vehicle',     # 大车
        1: 'vehicle',     # 小车
        6: 'vehicle',     # 特种车
        4: 'vehicle',     # 二轮车
        2: 'pedestrian',  # 行人 / 法规假人
    }


class BugScanProcessor:
    @staticmethod
    def get_BugScan_result(txt_path: str):
        txt_file = Path(txt_path)
        if not txt_file.exists():
            return None

        records = {}
        with open(txt_file, "r", encoding="utf-8", errors="ignore") as f:
            for line_num, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                    filename = data.get("文件名")
                    start_frame = data.get("开始帧")
                    end_frame = data.get("结束帧")
                    target_id = data.get("ARC_ID")
                    if filename is None or start_frame is None or end_frame is None or target_id is None:
                        continue
                    records.setdefault(filename, []).append(data)

                except json.JSONDecodeError as e:
                    print(f"⚠️ 第 {line_num} 行解析失败: {e}")
                    continue

        if not records:
            print("❌ 没有有效数据，退出")
            return None

        return records


class LabelInfoProcessor:
    """标签信息处理类"""
    
    @staticmethod
    def get_special_project_labelinfo(input_path):
        if os.path.splitext(input_path)[1] == ".xlsx":
            return LabelInfoProcessor._process_excel_file(input_path)
        else:
            raise ValueError(f"不支持的文件格式：{input_path}")
        
    @staticmethod
    def _process_excel_file(input_path):
        """处理Excel文件"""
        labelinfo_df = pd.read_excel(input_path, sheet_name="Sheet1", dtype=str)
        result = {}
        pre_scene = ""
        
        for index, row in labelinfo_df.iterrows():
            case_name = str(row.get("case_name", "")).strip()
            cur_scene = str(row.get("scene", "")).strip()
            
            # 大类继承
            if not cur_scene or cur_scene == "nan":
                cur_scene = pre_scene
            else:
                pre_scene = cur_scene

            filename = str(row.get("filename", "")).strip()
            video_dir = str(row.get("video_dir", "")).strip()
            start = row.get("start_frame", "")
            end = row.get("end_frame", "")
            target_id = str(row.get("target_id", "")).strip()
            class_key = row.get("class_key", None)
            class_value = row.get("class_value", None)
            
            # 缺少必要字段，跳过
            if not filename or not start or not end:
                continue
            
            # 转为整数
            try:
                start_frame = int(float(start))
                end_frame = int(float(end))
            except ValueError:
                print("⚠️ 跳过无效行，因为帧编号不是数字:", row)
                continue

            # 聚合逻辑
            if case_name not in result:
                result[case_name] = []

            try:
                target_id = int(target_id)
            except ValueError:
                pass
                
            result[case_name].append({
                "line_num": index + 1,
                "case_name": case_name,
                "scene": cur_scene,
                "filename": filename,
                "video_dir": video_dir,
                "start_frame": start_frame,
                "end_frame": end_frame,
                "target_id_list": [target_id],
                "class_key": class_key,
                "class_value": class_value
            })
        
        return result
    
    @staticmethod
    def find_specified_files(input_path, file_extensions=(".h264", ".h265")):
        search_results = {}
        if os.path.isdir(input_path):
            for root, _, files in os.walk(input_path):
                for file in files:
                    if file.lower().endswith(tuple(file_extensions)):
                        filename = os.path.splitext(file)[0]
                        if filename not in search_results:
                            search_results[filename] = os.path.join(root, file)
        elif os.path.isfile(input_path):
            with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
                for idx, line in enumerate(f, start=1):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        filename = os.path.splitext(os.path.basename(line))[0]
                        search_results[filename] = line
                    except json.JSONDecodeError:
                        print(f"⚠️ 第{idx}行不是合法 JSON: {line}")
        else:
            print("素材路径不存在")
        return search_results


class CSVFileHandler:
    """CSV文件处理类"""
    @staticmethod
    def find_csv_files(directory, key=".csv"):
        csv_files = []
        for file_name in os.listdir(directory):
            if file_name.endswith(key):
                csv_files.append(os.path.join(directory, file_name))
        return csv_files
    
    @staticmethod
    def process_all_csv_data(csv_files):
        """
        汇总目录下所有符合条件的 CSV 文件内容到一个 DataFrame 中
        """
        if not csv_files:
            print(f"未找到符合条件的文件。")
            return pd.DataFrame()

        # 汇总所有 CSV 文件内容
        combined_data = pd.DataFrame()
        for csv_file in csv_files:
            print(f"正在读取文件: {csv_file}")
            try:
                data = pd.read_csv(csv_file)
                combined_data = pd.concat([combined_data, data], ignore_index=True)
            except Exception as e:
                print(f"文件 {csv_file} 读取失败: {e}")

        return combined_data


class DataAggregator:
    """数据聚合处理类"""
    
    @staticmethod
    def safe_nan_func(group, func, default=np.nan):
        group = group[~np.isnan(group)]  # 移除 NaN 值
        if len(group) == 0:  # 如果全是 NaN 或数组为空
            return default
        return func(group)
    
    @staticmethod
    def get_percentile(group, percentile):
        """
        计算百分位数
        """
        # 排除 NaN 值
        group = group[~np.isnan(group)]
        if len(group) == 0:  # 如果全是 NaN
            return np.nan
        sorted_group = np.sort(group)
        index = int(np.ceil(len(sorted_group) * percentile)) - 1
        return sorted_group[index]


class ExcelFormatter:
    """Excel格式处理类"""
    
    @staticmethod
    def merge_same_cells(ws, df, col_name, col_index):
        prev_value = None
        start_index = 2  # Excel从第2行开始写数据

        for row_index, value in enumerate(df[col_name], start=2):
            # 判断是否值发生变化
            if value != prev_value:
                if prev_value is not None:
                    # 如果是不同的值，合并上一个区域
                    ws.merge_cells(
                        start_row=start_index,
                        end_row=row_index - 1,
                        start_column=col_index,
                        end_column=col_index
                    )
                    ws.cell(row=start_index, column=col_index).alignment = Alignment(
                        horizontal='center', vertical='center'
                    )

                # 更新起始位置
                start_index = row_index
                prev_value = value

        # 最后一个区域的合并
        ws.merge_cells(
            start_row=start_index,
            end_row=len(df) + 1,
            start_column=col_index,
            end_column=col_index
        )
        ws.cell(row=start_index, column=col_index).alignment = Alignment(
            horizontal='center', vertical='center'
        )
    
    @staticmethod
    def format_diff_columns(ws, df, diff_cols):
        diff_col_indices = [df.columns.get_loc(col) + 1 for col in diff_cols]
        diff_format = '[Red]"↑"0.000;[Green]"↓"0.000;0.000'

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in diff_col_indices:
                cell = row[col_idx - 1]  # openpyxl 的列是从0开始的
                cell.number_format = diff_format
    
    @staticmethod
    def center_align_cells(ws):
        """设置所有单元格居中且自动换行"""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
    
    @staticmethod
    def write_dataframe_to_sheet(ws, df, startrow=0):
        """将DataFrame写入工作表"""
        for r in dataframe_to_rows(df, index=False, header=(startrow == 0)):
            ws.append(r)


class SpeedIntervalProcessor:
    """速度区间处理类"""
    @staticmethod
    def get_speed_order_for_class(class_val):
        """根据class值获取对应的速度区间顺序"""
        class_str = str(class_val)
        
        # 检查数字类
        try:
            class_num = int(float(class_str))
            if class_num in DataProcessorConfig.VEHICLE_CLASS_SET:
                return DataProcessorConfig.VEHICLE_SPEED_ORDER
            elif class_num in DataProcessorConfig.PEDESTRIAN_CLASS_SET:
                return DataProcessorConfig.PEDESTRIAN_SPEED_ORDER
        except (ValueError, TypeError):
            pass
        
        # 检查字符串类
        if class_str in DataProcessorConfig.VEHICLE_CLASS_STR:
            return DataProcessorConfig.VEHICLE_SPEED_ORDER
        elif class_str in DataProcessorConfig.PEDESTRIAN_CLASS_STR:
            return DataProcessorConfig.PEDESTRIAN_SPEED_ORDER
        
        # 默认返回车辆速度顺序
        return DataProcessorConfig.VEHICLE_SPEED_ORDER
    
    @staticmethod
    def create_speed_intervals(df, target_config):
        """创建速度区间"""
        bins = target_config['bins']
        end_limit = target_config['end_limit']
        labels = [f'{bins[i]}-{bins[i + 1]}' for i in range(len(bins) - 1)]
        
        df = df.copy()
        # 将速度从m/s转换为km/h
        df['纵向速度区间'] = pd.cut(df['r_vx'] * 3.6, bins=bins, labels=labels, right=False)
        df['纵向速度区间_total'] = np.where(
            (df['r_vx'] * 3.6 >= 0) & (df['r_vx'] * 3.6 < end_limit),
            f'0-{end_limit}',
            '其他'
        )
        return df


class GroupedDataProcessor:
    """分组数据处理类"""
    def __init__(self, config):
        self.config = config
    
    def process_grouped_data(self, grouped, field, end_limit):
        # 对当前字段进行聚合
        field_new = field.split('_')[0]
        
        # 基础聚合
        result = self._perform_basic_aggregation(grouped, field, field_new)
        
        # 计算百分比
        result = self._calculate_percentages(result)
        
        # 计算百分位数
        result = self._calculate_percentiles(grouped, result, field)

        # 添加KPI和KPI达标率计算
        result = self._calculate_kpi_and_compliance(grouped, result, field)
        
        return result
    
    def _perform_basic_aggregation(self, grouped, field, field_new):
        """执行基础聚合"""
        use_percent = False
        if field == 'x_offset':
            field = 'x_offset_pre'
            use_percent = True

        return grouped.agg(
            **{
                'arc_数量': (field, lambda x: x.count()),
                'arc_均值': (f'r_{field_new}', lambda x: (
                    round(DataAggregator.safe_nan_func(x.abs(), np.nanmean, default=np.nan), 4)
                )),

                'arc_误差平均值': (field, lambda x: (
                    f"{round(DataAggregator.safe_nan_func(x, np.nanmean, default=np.nan) * 100, 2)}%"
                    if use_percent else
                    round(DataAggregator.safe_nan_func(x, np.nanmean, default=np.nan), 4)
                )),

                'arc_误差最小值': (field, lambda x: (
                    f"{round(DataAggregator.safe_nan_func(x, np.nanmin, default=np.nan) * 100, 2)}%"
                    if use_percent else
                    round(DataAggregator.safe_nan_func(x, np.nanmin, default=np.nan), 4)
                )),

                'arc_误差最大值': (field, lambda x: (
                    f"{round(DataAggregator.safe_nan_func(x, np.nanmax, default=np.nan) * 100, 2)}%"
                    if use_percent else
                    round(DataAggregator.safe_nan_func(x, np.nanmax, default=np.nan), 4)
                )),

                'arc_误差标准差': (field, lambda x: (
                    f"{round(DataAggregator.safe_nan_func(x, np.nanstd, default=np.nan) * 100, 2)}%"
                    if use_percent else
                    round(DataAggregator.safe_nan_func(x, np.nanstd, default=np.nan), 4)
                )),
            }
        ).reset_index()

    def _calculate_percentages(self, result):
        """计算百分比（兼容字符串 %）"""

        def to_float(series):
            return pd.to_numeric(
                series.astype(str).str.replace('%', '', regex=False),
                errors='coerce'
            )

        err = to_float(result['arc_误差平均值'])
        mean = to_float(result['arc_均值'])

        result['arc_误差占比百分比'] = np.where(
            (~np.isnan(err)) &
            (~np.isnan(mean)) &
            (mean != 0),
            (err / mean * 100).round(4).astype(str) + '%',
            np.nan
        )

        return result

    
    def _calculate_percentiles(self, grouped, result, field):
        """计算百分位数"""
        # 计算百分位数
        use_percent = False
        if field == 'x_offset':
            field = 'x_offset_pre'
            use_percent = True

        for current_field in [field]:
            percentile_50 = grouped[current_field].apply(lambda x: DataAggregator.get_percentile(x, 0.5))
            percentile_90 = grouped[current_field].apply(lambda x: DataAggregator.get_percentile(x, 0.9))
            percentile_95 = grouped[current_field].apply(lambda x: DataAggregator.get_percentile(x, 0.95))

            # 构造当前字段的结果 DataFrame
            key = 'arc'
            def format_value(val):
                # NaN / None → 空
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    return None   # 或者 ''，看你后面 Excel 想不想留空

                if use_percent:
                    return f"{round(val * 100, 2)}%"
                return round(val, 4)

            result[f'{key}_percentile:0.5'] = [format_value(v) for v in percentile_50.values]
            result[f'{key}_percentile:0.9'] = [format_value(v) for v in percentile_90.values]
            result[f'{key}_percentile:0.95'] = [format_value(v) for v in percentile_95.values]

        return result
    
    def _calculate_kpi_and_compliance(self, grouped, result, field):
        """计算KPI和KPI达标率 - 基于每行的达标状态计算"""
        try:
            # 获取当前字段对应的KPI配置键
            kpi_key = DataProcessorConfig.FIELD_TO_KPI_KEY.get(field)
            if not kpi_key:
                return result
            
            # 获取对应的状态列名和KPI列名
            status_col = DataProcessorConfig.FIELD_TO_STATUS_COL.get(field)
            kpi_col = DataProcessorConfig.FIELD_TO_KPI_COL.get(field)
            
            if not status_col or not kpi_col:
                return result
            
            # 创建一个字典来存储每个分组的KPI结果
            kpi_dict = {}
            
            # 首先获取所有分组键
            all_groups = list(grouped.groups.keys())
            
            for group_key in all_groups:
                try:
                    # 获取分组数据
                    group_df = grouped.get_group(group_key)
                    class_val, category, speed_range = group_key
                    
                    # 获取分组中有效的达标状态数据
                    if status_col in group_df.columns:
                        # 过滤掉状态为NaN的行
                        valid_mask = group_df[status_col].notna()
                        valid_status = group_df.loc[valid_mask, status_col]
                        
                        if len(valid_status) == 0:
                            kpi_dict[group_key] = {
                                'KPI需求值': np.nan,
                                'KPI达标率': np.nan
                            }
                            continue
                        
                        # 计算达标率（True状态的占比）
                        compliant_count = valid_status.sum()  # True会被视为1，False为0
                        total_count = len(valid_status)
                        compliance_rate = compliant_count / total_count if total_count > 0 else 0
                        
                        # 获取该分组的KPI需求值（取第一个非NaN值）
                        if kpi_col in group_df.columns:
                            # 过滤掉KPI需求值为NaN的行
                            kpi_mask = group_df[kpi_col].notna()
                            if kpi_mask.any():
                                # 取第一个有效的KPI需求值
                                kpi_value = group_df.loc[kpi_mask, kpi_col].iloc[0]
                                
                                # 对于纵向距离KPI，格式化为百分比
                                if kpi_key == '纵向距离' and pd.notna(kpi_value):
                                    try:
                                        # 如果kpi_value已经是字符串，直接使用
                                        if isinstance(kpi_value, str):
                                            kpi_threshold_val = kpi_value
                                        else:
                                            # 否则格式化为百分比
                                            kpi_threshold_val = f"{float(kpi_value)}%"
                                    except:
                                        kpi_threshold_val = f"{kpi_value}%"
                                else:
                                    kpi_threshold_val = kpi_value
                            else:
                                kpi_threshold_val = np.nan
                        else:
                            kpi_threshold_val = np.nan
                        
                        # 如果class_val是"ALL"，将KPI需求值设为NaN
                        class_val_str = str(class_val)
                        if class_val_str in ["ALL"]:
                            kpi_threshold_val = np.nan
                        
                        kpi_dict[group_key] = {
                            'KPI需求值': kpi_threshold_val,
                            'KPI达标率': compliance_rate
                        }
                    else:
                        kpi_dict[group_key] = {
                            'KPI需求值': np.nan,
                            'KPI达标率': np.nan
                        }
                        
                except Exception as e:
                    print(f"计算分组 {group_key} 的KPI时出错: {e}")
                    import traceback
                    traceback.print_exc()
                    kpi_dict[group_key] = {
                        'KPI需求值': np.nan,
                        'KPI达标率': np.nan
                    }
            
            # 将KPI结果合并到result中
            kpi_results = []
            compliance_results = []
            
            # 遍历result中的每一行，查找对应的KPI结果
            for idx, row in result.iterrows():
                class_val = row['class']
                category = row['category']
                speed_range = row.get('纵向速度区间', row.get('纵向速度区间_total', ''))
                
                group_key = (class_val, category, speed_range)
                
                if group_key in kpi_dict:
                    if speed_range == '0-200' or row.get('纵向速度区间', None) is None:
                        kpi_results.append(np.nan)
                    else:
                        kpi_results.append(kpi_dict[group_key]['KPI需求值'])
                    compliance_results.append(kpi_dict[group_key]['KPI达标率'])
                else:
                    kpi_results.append(np.nan)
                    compliance_results.append(np.nan)
            
            # 将结果添加到DataFrame中
            result['KPI需求值'] = kpi_results
            result['KPI达标率'] = compliance_results
            
            # 格式化达标率为百分比
            result['KPI达标率'] = result['KPI达标率'].apply(
                lambda x: f"{x*100:.2f}%" if pd.notna(x) else np.nan
            )
            
        except Exception as e:
            print(f"计算KPI时出错: {e}")
            import traceback
            traceback.print_exc()
        
        return result

    def _parse_kpi_threshold(self, threshold_str, group_df, field):
        """解析KPI阈值字符串"""
        if isinstance(threshold_str, (int, float)):
            return float(threshold_str)
        
        threshold_str = str(threshold_str).strip()
        
        # 处理 max(10%, 1) 这种格式
        if threshold_str.startswith('max(') and threshold_str.endswith(')'):
            inner = threshold_str[4:-1]
            parts = inner.split(',')
            
            thresholds = []
            for part in parts:
                part = part.strip()
                if '%' in part:
                    # 百分比阈值，需要计算相对值
                    percent = float(part.replace('%', '')) / 100
                    
                    # 获取真实值列名
                    if field == 'vx_offset':
                        real_col = 'r_vx'
                    elif field == 'vy_offset':
                        real_col = 'r_vy'
                    else:
                        # 对于距离字段，使用绝对阈值
                        real_col = None
                    
                    if real_col and real_col in group_df.columns:
                        # 计算百分比阈值（相对误差）
                        real_values = group_df[real_col].abs()
                        if len(real_values) > 0:
                            threshold = percent * real_values.mean()
                        else:
                            threshold = 0
                    else:
                        threshold = percent
                else:
                    # 绝对值阈值
                    threshold = float(part)
                
                thresholds.append(threshold)
            
            # 返回最大值
            return max(thresholds)
        
        # 如果是纯数字字符串
        try:
            return float(threshold_str)
        except ValueError:
            return np.nan

    def _calculate_compliance_rate(self, data, kpi_threshold, field, kpi_key):
        """计算达标率"""
        if np.isnan(kpi_threshold) or len(data) == 0:
            return np.nan
        
        if kpi_key in ['纵向距离', '横向距离', '纵向速度', '横向速度']:
            # 对于这些指标，达标条件是误差绝对值 <= KPI阈值
            compliant_count = np.sum(np.abs(data) <= kpi_threshold)
            return compliant_count / len(data)
        else:
            return np.nan


class VideoErrorProcessor:
    """视频误差处理类"""
    
    @staticmethod
    def fmt_pct(val, digits=2):
        """数值 -> 'xx.xx%'，None / nan 原样返回"""
        if val is None or pd.isna(val):
            return None
        return f"{val:.{digits}f}%"

    @staticmethod
    def get_video_error_result(df, config):
        end_limit = config['end_limit']
        bins = config['bins']

        # 1️⃣ 距离过滤
        df = df[((df['CIPV'] == 1) & (df['x'] >= 5)) | ((df['CIPV'] != 1) & (df['x'] >= 10))].copy()

        df = df[(df['r_x'] >= 0) & (df['r_x'] <= end_limit)]
        if df.empty:
            return None

        # 2️⃣ 速度区间（使用km/h）
        bin_labels = [f"{bins[i]}-{bins[i+1]}" for i in range(len(bins)-1)]
        df['speed_bin'] = pd.cut(df['r_vx'] * 3.6, bins=bins, labels=bin_labels, right=False)

        results = []

        for bin_label, grp in df.groupby('speed_bin'):
            if grp.empty:
                continue

            row = {
                '纵向速度区间': bin_label,
                '样本点数': len(grp),
            }

            # ================= 纵向距离（百分比） =================
            x_pct_err = (grp['x_offset'].abs() / grp['x'].abs().clip(1e-6)) * 100
            kpi_x = DataProcessorConfig.KPI_CONFIGS.get('vehicle', {}).get('纵向距离', {}).get(bin_label, None)
            if kpi_x is not None:
                row['纵向距离需求KPI'] = f"{kpi_x}%"
                row['纵向距离达标率'] = VideoErrorProcessor.fmt_pct(
                    (x_pct_err <= kpi_x).mean() * 100
                )
            else:
                row['纵向距离需求KPI'] = None
                row['纵向距离达标率'] = None

            row['纵向距离平均误差'] = VideoErrorProcessor.fmt_pct(x_pct_err.mean())
            row['纵向距离P50'] = VideoErrorProcessor.fmt_pct(x_pct_err.quantile(0.50))
            row['纵向距离P90'] = VideoErrorProcessor.fmt_pct(x_pct_err.quantile(0.90))
            row['纵向距离P95'] = VideoErrorProcessor.fmt_pct(x_pct_err.quantile(0.95))

            
            # ================= 横向距离 =================
            y_abs = grp['y_offset'].abs()
            kpi_y = DataProcessorConfig.KPI_CONFIGS.get('vehicle', {}).get('横向距离', {}).get(bin_label, None)
            if kpi_y is not None:
                row['横向距离需求KPI'] = kpi_y
                row['横向距离达标率'] = VideoErrorProcessor.fmt_pct(
                    (y_abs <= kpi_y).mean() * 100
                )
            else:
                row['横向距离需求KPI'] = None
                row['横向距离达标率'] = None

            row['横向距离平均误差'] = y_abs.mean()
            row['横向距离P50'] = y_abs.quantile(0.50)
            row['横向距离P90'] = y_abs.quantile(0.90)
            row['横向距离P95'] = y_abs.quantile(0.95)

            # ================= 纵向速度 =================
            vx_err = grp['vx_offset'].abs()   # m/s
            vx_abs = grp['vx'].abs().clip(1e-6)
            vx_pct = (grp['vx_offset'].abs() / vx_abs) * 100

            kpi_vx = DataProcessorConfig.KPI_CONFIGS.get('vehicle', {}).get('纵向速度', {}).get(bin_label, None)
            if kpi_vx == 'max(10%, 1)':
                vx_kpi_val = np.maximum(10, 1 / vx_abs * 100)
                row['纵向速度需求KPI'] = 'max(10%, 1)'
                row['纵向速度达标率'] = VideoErrorProcessor.fmt_pct(
                    (vx_pct <= vx_kpi_val).mean() * 100
                )
            else:
                row['纵向速度需求KPI'] = None
                row['纵向速度达标率'] = None

            row['纵向速度平均误差'] = vx_err.mean()
            row['纵向速度P50'] = vx_err.quantile(0.50)
            row['纵向速度P90'] = vx_err.quantile(0.90)
            row['纵向速度P95'] = vx_err.quantile(0.95)

            # ================= 横向速度 =================
            vy_err = grp['vy_offset'].abs()
            vy_abs = grp['vy'].abs().clip(1e-6)
            vy_pct = (grp['vy_offset'].abs() / vy_abs) * 100

            kpi_vy = DataProcessorConfig.KPI_CONFIGS.get('vehicle', {}).get('横向速度', {}).get(bin_label, None)
            if kpi_vy == 'max(10%, 1)':
                vy_kpi_val = np.maximum(10, 1 / vy_abs * 100)
                row['横向速度需求KPI'] = 'max(10%, 1)'
                row['横向速度达标率'] = VideoErrorProcessor.fmt_pct(
                    (vy_pct <= vy_kpi_val).mean() * 100
                )
            else:
                row['横向速度需求KPI'] = None
                row['横向速度达标率'] = None

            row['横向速度平均误差'] = vy_err.mean()
            row['横向速度P50'] = vy_err.quantile(0.50)
            row['横向速度P90'] = vy_err.quantile(0.90)
            row['横向速度P95'] = vy_err.quantile(0.95)

            results.append(row)

        return pd.DataFrame(results)


class ARCV3DataProcessor:
    """主数据处理类"""
    
    def __init__(self, config=None, replay_txtlist=None):
        self.config = config or DataProcessorConfig.UNIFIED_CONFIG
        self.grouped_processor = GroupedDataProcessor(self.config)
        self.replay_txtlist = replay_txtlist

        self.bugscan_target_result = {}
        self.scene_class_target_counts = {}
    
    def process_data(self, input_dir, output_file, labelinfo, bugscan_result=None, key=".csv"):
        # 查找符合条件的 CSV 文件
        csv_files = CSVFileHandler.find_csv_files(input_dir, key=key)
        
        # 处理所有CSV文件
        df = CSVFileHandler.process_all_csv_data(csv_files)
        
        # 数据预处理
        df_processed = self._preprocess_data(df)        
        if df_processed.empty:
            print("处理后的数据为空，无法继续处理")
            return
        
        # 计算误差
        df_with_errors = self._calculate_errors(df_processed)
        case_filtered_dfs = {}

        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            scene_class_target_counts = {}
            for case_name, label_attr in labelinfo.items():
                for label_data in label_attr:
                    filename = label_data.get("filename")
                    case_name = label_data.get("case_name")
                    video_dir = label_data.get("video_dir")
                    scene = label_data.get("scene")
                    start_frame = label_data.get("start_frame")
                    end_frame = label_data.get("end_frame")
                    target_id_list = label_data.get("target_id_list", [])
                    class_key = label_data.get("class_key", None)
                    class_value = label_data.get("class_value", None)
                    if end_frame - start_frame <= 10:
                        continue # 过滤过短的场景（小于10帧）

                    df_file = df_with_errors[df_with_errors["filename"] == f"{filename}.txt"]  # 先按 filename 粗过滤一层，减少数据量
                    # 条件过滤
                    matched_df = df_file[
                        (df_file["frame_index"] >= start_frame) &
                        (df_file["frame_index"] <= end_frame) &
                        (df_file["reVeh_tid"].isin(target_id_list))
                    ]

                    # matched_df 匹配专项的结果
                    if not matched_df.empty:
                        # ============================================================
                        # 如果标注得有class_key和class_value，修改matched_df中的class
                        obj_class = self._process_class_key(class_key, matched_df)
                        if obj_class is None:
                            continue
                        # ============================================================

                        # 分别处理车辆和行人数据
                        df_processed = self._process_vehicle_pedestrian_data(matched_df)
                        
                        # ============================================================
                        # bug扫描匹配结果
                        tid_list = df_processed['tid'].unique().tolist() # 获取matched_df中tid列的所有值（去除重复）
                        rid_list = df_processed['rid'].unique().tolist() # 获取matched_df中tid列的所有值（去除重复）
                        self._process_bugscan_result(filename, bugscan_result, start_frame, end_frame, 
                                                    tid_list, case_name, scene, obj_class)
                        # ============================================================
                        df_processed['case_name'] = case_name # 添加 case_name 列
                        df_processed['场景'] = scene    # 添加 scene 列

                        # 计算KPI相关列
                        df_processed = self._calculate_kpi_columns(df_processed)

                        # 用于计算每个case的汇总表
                        if case_name not in case_filtered_dfs:
                            case_filtered_dfs[case_name] = {}
                        
                        if scene not in case_filtered_dfs[case_name]:
                            case_filtered_dfs[case_name][scene] = []
                        
                        case_filtered_dfs[case_name][scene].append(df_processed)

                        # ============================================================
                        # 用于添加汇总中目标的数量（不是单帧）
                        self._update_scene_class_target_counts(case_name, scene, filename, start_frame, 
                                                              end_frame, target_id_list, df_processed)
                        
                    else:
                        # 未匹配到实例专项的目标，初始化默认值
                        df_processed = pd.DataFrame()
                        tid_list = []
                        rid_list = []
                        attr = {
                            "case_name": case_name,
                            "scene": scene,
                            "filename": filename,
                            "start_frame": start_frame,
                            "end_frame": end_frame,
                            "target_id_list": target_id_list,
                            "class_key": class_key
                        }
                        print(f"[未匹配到] {attr}")
                    
                    self._process_Singleinstance(df_processed, filename, scene, class_key, class_value, video_dir, 
                                               start_frame, end_frame, target_id_list, tid_list, rid_list, writer)

            # ================================================
            # 将集合转换为计数（用于计算目标的数量）
            self._convert_target_counts_to_numbers(scene_class_target_counts)
            # ===============================================
            
            # 计算每个case的每个scene的汇总
            case_scene_summaries = self._create_case_scene_summaries(case_filtered_dfs)
            
            # 写入Excel
            self._write_to_excel(case_scene_summaries, writer)
    
    def _process_class_key(self, class_key, matched_df):
        """处理class_key"""
        if class_key:
            if class_key in ["commercial_vehicle_types", "engineering_vehicle_types"]: # 大车
                matched_df['class'] = 0
                return 0
            elif class_key in ["small_vehicle_types"]: # 小车
                matched_df['class'] = 1
                return 1
            elif class_key in ["pedestrian_types", "regulatory_targets"]: # 行人, 法规假人
                matched_df['class'] = 2
                return 2
            elif class_key in ["two_wheeled_vehicle_types"]: # 骑行人
                matched_df['class'] = 4
                return 4
            elif class_key in ["special_vehicle_types"]: # 特种车
                matched_df['class'] = 6
                return 6
            else:
                print("不支持的class_key:", class_key)
                return None
        else:
            # 使用算法输出的class作为目标类型计算
            return None
    
    def _process_vehicle_pedestrian_data(self, matched_df):
        """处理车辆和行人数据"""
        speed_config = self.config['speed_config']
        
        # 为所有数据创建速度区间
        df_processed = SpeedIntervalProcessor.create_speed_intervals(matched_df, speed_config)
        
        return df_processed
    
    def _process_bugscan_result(self, filename, bugscan_result, start_frame, end_frame, 
                               tid_list, case_name, scene, obj_class):
        """处理bugscan结果"""
        bugResult_file = bugscan_result.get(filename, None)
        case_passflag = {"纵向测距": True, "横向测距": True, "纵向测速": True, "横向测速": True}
        if bugResult_file is None:
            return
        
        for bugResult in bugResult_file:
            signal_type = bugResult.get("感知信号")
            start_f = bugResult.get("开始帧")
            end_f = bugResult.get("结束帧")
            arc_ID = bugResult.get("ARC_ID")
            if (start_frame <= start_f <= end_frame or start_frame <= end_f <= end_frame) and arc_ID in tid_list:
                if case_name not in self.bugscan_target_result:
                    self.bugscan_target_result[case_name] = {}
                if scene not in self.bugscan_target_result[case_name]:
                    self.bugscan_target_result[case_name][scene] = {}
                if obj_class not in self.bugscan_target_result[case_name][scene]:
                    self.bugscan_target_result[case_name][scene][obj_class] = {
                        "纵向测距": 0,
                        "横向测距": 0,
                        "纵向测速": 0,
                        "横向测速": 0,
                    }
                if signal_type == "纵向测距" and case_passflag["纵向测距"]:
                    case_passflag["纵向测距"] = False
                    self.bugscan_target_result[case_name][scene][obj_class]["纵向测距"] += 1
                elif signal_type == "横向测距" and case_passflag["横向测距"]:
                    case_passflag["横向测距"] = False
                    self.bugscan_target_result[case_name][scene][obj_class]["横向测距"] += 1
                elif signal_type == "纵向测速" and case_passflag["纵向测速"]:
                    case_passflag["纵向测速"] = False
                    self.bugscan_target_result[case_name][scene][obj_class]["纵向测速"] += 1
                elif signal_type == "横向测速" and case_passflag["横向测速"]:
                    case_passflag["横向测速"] = False
                    self.bugscan_target_result[case_name][scene][obj_class]["横向测速"] += 1
    
    def _calculate_kpi_columns(self, df_processed):
        """计算KPI相关列"""
        # 注意：KPI配置现在基于速度区间
        df_processed['纵向距离_KPI需求值'] = df_processed.apply(lambda row: self.get_kpi_threshold(row, '纵向距离'), axis=1)
        df_processed['纵向距离_KPI达标状态'] = ((df_processed['x'] - df_processed['r_x']).abs() / df_processed['r_x']) * 100 <= df_processed['纵向距离_KPI需求值']
        df_processed['横向距离_KPI需求值'] = df_processed.apply(lambda row: self.get_kpi_threshold(row, '横向距离'), axis=1)
        df_processed['横向距离_KPI达标状态'] = (df_processed['y'] - df_processed['r_y']).abs() <= df_processed['横向距离_KPI需求值']

        df_processed['纵向速度_KPI需求值'] = df_processed.apply(lambda row: self.get_kpi_threshold(row, '纵向速度'), axis=1)
        df_processed['纵向速度_KPI达标状态'] = (df_processed['vx'] - df_processed['r_vx']).abs() <= np.maximum(0.1 * df_processed['r_vx'].abs(), 1.0)

        df_processed['横向速度_KPI需求值'] = df_processed.apply(lambda row: self.get_kpi_threshold(row, '横向速度'), axis=1)
        df_processed['横向速度_KPI达标状态'] = (df_processed['vy'] - df_processed['r_vy']).abs() <= np.maximum(0.1 * df_processed['r_vy'].abs(), 1.0)
        
        return df_processed
    
    def _update_scene_class_target_counts(self, case_name, scene, filename, start_frame, 
                                         end_frame, target_id_list, df_processed):
        """更新场景类目标计数"""
        if case_name not in self.scene_class_target_counts:
            self.scene_class_target_counts[case_name] = {}
        if scene not in self.scene_class_target_counts[case_name]:
            self.scene_class_target_counts[case_name][scene] = {}

        # 为每个目标ID分别统计class
        for target_id in target_id_list:
            # 获取这个目标ID对应的数据
            target_df = df_processed[df_processed["reVeh_tid"] == target_id]
            if not target_df.empty:
                # 获取这个目标的主要class（取第一个或最常见的class）
                if 'class' in target_df.columns:
                    class_counts = target_df['class'].value_counts()
                    if not class_counts.empty:
                        main_class = str(class_counts.index[0])
                        # 构建唯一标识：使用目标ID来区分同一场景中的不同目标
                        target_key = f"{filename}_{start_frame}_{end_frame}_{target_id}_{main_class}"
                        if main_class not in self.scene_class_target_counts[case_name][scene]:
                            self.scene_class_target_counts[case_name][scene][main_class] = set()
                        # 添加唯一的目标标识
                        self.scene_class_target_counts[case_name][scene][main_class].add(target_key)
    
    def _convert_target_counts_to_numbers(self, scene_class_target_counts):
        """将目标计数集合转换为数字"""
        for case_name in self.scene_class_target_counts:
            for scene in self.scene_class_target_counts[case_name]:
                for class_key in self.scene_class_target_counts[case_name][scene]:
                    self.scene_class_target_counts[case_name][scene][class_key] = len(
                        self.scene_class_target_counts[case_name][scene][class_key]
                    )
    
    def _create_case_scene_summaries(self, case_filtered_dfs):
        """创建场景汇总"""
        case_scene_summaries = {}
        for case_name, scene_dict in case_filtered_dfs.items():
            case_scene_summaries[case_name] = {}
            for scene, data_list in scene_dict.items():
                if data_list:
                    # 合并scene数据
                    scene_df = pd.concat(data_list, ignore_index=True)
                    case_scene_summaries[case_name][scene] = scene_df
                else:
                    case_scene_summaries[case_name][scene] = pd.DataFrame()
        return case_scene_summaries

    def get_kpi_threshold(self, row, signal_type):
        """
        row: df 的一行
        signal_type: '纵向距离' / '横向距离' / '纵向速度' / '横向速度'
        """
        def parse_range(range_str):
            low, high = range_str.split('-')
            return float(low), float(high)

        def match_range(speed_kmh, range_dict):
            fallback_key = None

            for range_key in range_dict:
                low, high = parse_range(range_key)

                # 记录兜底区间（如 0-200）
                if low == 0 and high >= 200:
                    fallback_key = range_key

                if low <= speed_kmh < high:
                    return range_key

            return fallback_key
        
        class_id = row['class']   # 0 / 1 / 2 / 4 / 6
        speed_kmh = row['r_vx'] * 3.6  # 纵向速度（km/h）

        # 1. class → KPI 大类
        kpi_type = DataProcessorConfig.CLASS_TO_KPI_TYPE.get(class_id)
        if kpi_type is None:
            return None

        # 2. 获取该类 KPI 配置
        signal_config = DataProcessorConfig.KPI_CONFIGS[kpi_type].get(signal_type)
        if signal_config is None:
            return None

        # 3. 匹配速度区间
        range_key = match_range(speed_kmh, signal_config)
        if range_key is None:
            return None

        # 4. 返回阈值
        return signal_config[range_key]

    def _preprocess_data(self, df):
        """数据预处理"""
        if df.empty:
            return pd.DataFrame()
        
        # 过滤新雷达 999
        df = df[(df['r_vx'] != 999) & (df['r_vy'] != 999)]
        
        # 只对 class 为 2 或 4 的行进行 vx、vy 小于等于 0.2 的过滤
        mask_class = df['class'].isin([2, 4])
        mask_small_v = (df['r_vx'].abs() <= 0.2) & (df['r_vy'].abs() <= 0.2)
        df = df[~(mask_class & mask_small_v)]

        return df
        
    def _calculate_errors(self, df_processed):
        """计算误差"""
        col_list = ['x', 'y', 'vx', 'vy']
        for v in col_list:
            # 误差绝对值
            if "_" in v:
                r_v = v.split("_")[1]
            else:
                r_v = v
            
            # 误差绝对值
            df_processed[f'{v}_offset'] = abs(df_processed[f'{v}'] - df_processed[f'r_{r_v}'])
            
            # 误差百分比
            df_processed[f'{v}_offset_pre'] = abs(df_processed[f'{v}_offset'] / df_processed[f'r_{r_v}'])
        
        return df_processed
    
    # 处理单个实例的各个KPI
    def _process_Singleinstance(
        self, matched_df, filename,
        scene, class_key, class_value, video_dir, start_frame, end_frame,
        target_id_list, tid_list, rid_list, writer
    ):
        """处理单个 case 的视频误差"""
        if matched_df.empty:
            return

        all_config = self.config['all_config']

        # 1️⃣ 计算误差结果（多行：速度区间）
        video_df = VideoErrorProcessor.get_video_error_result(matched_df, all_config)
        if video_df is None or video_df.empty:
            return
        
        replay_txtpath = self.replay_txtlist.get(filename)

        # 2️⃣ 补充 case 公共字段
        video_df.insert(0, '场景', scene)
        video_df.insert(1, '目标类别', class_key)
        video_df.insert(2, '目标类别值', class_value)
        video_df.insert(3, '视频目录', video_dir)
        video_df.insert(4, '回灌路径', replay_txtpath)
        video_df.insert(5, '文件名', filename)
        video_df.insert(6, '标注目标ID', ','.join(map(str, target_id_list)))
        video_df.insert(7, '新版本目标ID', ','.join(map(str, tid_list)))
        video_df.insert(8, '雷达目标ID', ','.join(map(str, rid_list)))
        video_df.insert(9, '开始帧', start_frame)
        video_df.insert(10, '结束帧', end_frame)
        
        sheet_name = '每个case统计结果'

        # 3️⃣ 写入 Excel
        startrow = (
            writer.sheets[sheet_name].max_row
            if sheet_name in writer.sheets
            else 0
        )

        video_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=startrow,
            header=(startrow == 0),
        )

    def _write_to_excel(self, case_combined_data, writer):
        """
        按 scene 聚合，合并四个误差字段为一个 sheet
        """
        # 全局结果缓存（合并为一个表格）
        all_combined_tables = []
        case_all_combined_tables = []
        totle_combined_tables = []

        category_order = ["ALL", "CIPV", "三车道[-8,8]", "左车道[2,8]", "自车道[-2,2]", "右车道[-8,-2]"]

        # 对每个 scene 处理
        totle_combined_dfs = []
        for case_name, scene_combined_dict in case_combined_data.items():
            case_all_dfs = []
            for scene_name, scene_data in scene_combined_dict.items():
                if scene_data.empty:
                    continue
                case_all_dfs.append(scene_data) #合并所有case_name的数据
                totle_combined_dfs.append(scene_data)

                # 处理分组数据，获取合并后的表格
                combined_table = self._process_grouped_data(scene_data, scene_name, case_name, category_order)

                if combined_table is not None and not combined_table.empty:
                    all_combined_tables.append(combined_table)
            
            # 计算单个case的汇总
            case_conbined_single_df = pd.concat(case_all_dfs, ignore_index=True)
            case_combined_table = self._process_case_grouped_data(case_conbined_single_df, case_name, category_order)
            if case_combined_table is not None and not case_combined_table.empty:
                case_all_combined_tables.append(case_combined_table)
        
        # 计算所有汇总的
        totle_combined_df = pd.concat(totle_combined_dfs, ignore_index=True)
        totle_combined_table = self._process_totle_grouped_data(totle_combined_df, category_order)
        if totle_combined_table is not None and not totle_combined_table.empty:
            totle_combined_tables.append(totle_combined_table)
        
        # 合并所有表格
        all_final_combined_df = pd.concat(totle_combined_tables, ignore_index=True) if totle_combined_tables else pd.DataFrame()
        final_combined_df = pd.concat(all_combined_tables, ignore_index=True) if all_combined_tables else pd.DataFrame()
        case_final_combined_df = pd.concat(case_all_combined_tables, ignore_index=True) if case_all_combined_tables else pd.DataFrame()

        # 写入合并后的sheet
        self._write_combined_result_to_excel(final_combined_df, writer, '结果汇总')
        self._write_combined_result_to_excel(case_final_combined_df, writer, 'case结果汇总')
        self._write_combined_result_to_excel(all_final_combined_df, writer, 'all结果汇总', is_all=True)

    def _process_case_grouped_data(self, df_processed, case_name, category_order):
        """
        单个 scene
        → 返回合并后的DataFrame（包含四个误差字段）
        """
        if df_processed.empty:
            return pd.DataFrame()
        
        data_frames = self._create_category_dataframes(df_processed)
        combined_df = pd.concat(list(data_frames.values()), ignore_index=True)

        if combined_df.empty:
            return pd.DataFrame()

        # 处理四个误差字段，返回合并后的表格
        combined_table = self._process_by_fields(combined_df, category_order)
        
        # 给表格加 scene 信息
        if combined_table is not None and not combined_table.empty:
            combined_table.insert(0, 'case_name', case_name)

        return combined_table
    
    def _process_totle_grouped_data(self, df_processed, category_order):
        """
        单个 scene
        → 返回合并后的DataFrame（包含四个误差字段）
        """
        if df_processed.empty:
            return pd.DataFrame()
        
        data_frames = self._create_category_dataframes(df_processed)
        combined_df = pd.concat(list(data_frames.values()), ignore_index=True)

        if combined_df.empty:
            return pd.DataFrame()

        # 处理四个误差字段，返回合并后的表格
        combined_table = self._process_by_fields(combined_df, category_order)

        return combined_table

    def _process_grouped_data(self, df_processed, scene_name, case_name, category_order):
        """
        单个 scene
        → 返回合并后的DataFrame（包含四个误差字段）
        """
        if df_processed.empty:
            return pd.DataFrame()
        
        data_frames = self._create_category_dataframes(df_processed)
        combined_df = pd.concat(list(data_frames.values()), ignore_index=True)

        if combined_df.empty:
            return pd.DataFrame()

        # 处理四个误差字段，返回合并后的表格
        combined_table = self._process_by_fields(combined_df, category_order)
        
        # 给表格加 scene 信息
        if combined_table is not None and not combined_table.empty:
            combined_table.insert(0, 'case_name', case_name)
            combined_table.insert(1, 'scene', scene_name)

        return combined_table

    def _create_category_dataframes(self, df_processed):
        """创建不同类别的数据框"""
        data_frames = {}
        # 添加 class/category 汇总的 ALL 行（用于生成每个 scene 的汇总）
        df_all = df_processed[
            (df_processed['y'].between(-8, 8)) &
            (
                ((df_processed['CIPV'] == 1) & (df_processed['x'] >= 5)) |
                ((df_processed['CIPV'] != 1) & (df_processed['x'] >= 10))
            )
        ].copy()
        df_all['class'] = 'ALL'
        df_all['category'] = 'ALL'
        data_frames['ALL'] = df_all
        
        # 表一：筛选 CIPV = 1
        df_cipv = df_processed[(df_processed['CIPV'] == 1) & (df_processed['x'] >= 5)].copy()
        df_cipv['category'] = 'CIPV'
        data_frames['cipv'] = df_cipv

        # 表二：筛选 Y轴范围 [-8,8]
        df_y_range = df_processed[
            (df_processed['y'].between(-8, 8)) &
            (
                ((df_processed['CIPV'] == 1) & (df_processed['x'] >= 5)) |
                ((df_processed['CIPV'] != 1) & (df_processed['x'] >= 10))
            )
        ].copy()

        df_y_range['category'] = '三车道[-8,8]'
        data_frames['y_range'] = df_y_range

        # 表三：左车道 Y轴范围 [2, 8]
        df_y_left = df_processed[
            (df_processed['y'] >= 2) & 
            (df_processed['y'] <= 8) & 
            (df_processed['x'] >= 10)
        ].copy()
        df_y_left['category'] = '左车道[2,8]'
        data_frames['y_left'] = df_y_left

        # 表四：自车道 Y轴范围 [-2, 2]
        df_y_current = df_processed[
            (df_processed['y'] >= -2) & 
            (df_processed['y'] <= 2) & 
            (df_processed['x'] >= 10)
        ].copy()
        df_y_current['category'] = '自车道[-2,2]'
        data_frames['y_current'] = df_y_current

        # 表五：右车道 Y轴范围 [-8, -2]
        df_y_right = df_processed[
            (df_processed['y'] >= -8) & 
            (df_processed['y'] <= -2) & 
            (df_processed['x'] >= 10)
        ].copy()
        df_y_right['category'] = '右车道[-8,-2]'
        data_frames['y_right'] = df_y_right
        
        return data_frames

    def _process_by_fields(self, combined_df, category_order):
        """
        按误差字段处理，返回合并后的表格
        → 返回合并后的DataFrame
        """
        x_list = ['纵向速度区间', '纵向速度区间_total']
        fields_list = {
            'x_offset': 'DX_纵向距离误差',
            'vx_offset': 'DVX_纵向速度误差',
            'y_offset': 'DY_横向距离误差',
            'vy_offset': 'DVY_横向速度误差'
        }

        classifications = self.config['classifications']
        field_results = {}

        for field, tag in fields_list.items():
            all_results = []

            for x in x_list:
                classification_result = []

                for classification, class_values in classifications.items():
                    if not class_values:
                        continue
                    
                    # 对于 ALL，只保留 total 分组
                    if classification == 'ALL':
                        if x == '纵向速度区间':
                            continue
                    
                    filtered_df = combined_df[combined_df['class'].isin(class_values)].copy()
                    if filtered_df.empty:
                        continue

                    filtered_df['class'] = classification
                    grouped = filtered_df.groupby(['class', 'category', x])

                    if classification == 'ALL':
                        end_limit = self.config['all_config']['end_limit']
                    elif any(cls in [0, 1, 3, 6] for cls in class_values):
                        end_limit = self.config['vehicle_config']['end_limit']
                    else:
                        end_limit = self.config['pedestrian_config']['end_limit']

                    class_result = self.grouped_processor.process_grouped_data(grouped, field, end_limit)
                    
                    if class_result is not None and not class_result.empty:
                        # 如果当前是按 total 分组（例如 '纵向速度区间_total'），
                        # 将该列重命名为统一的 '纵向速度区间'，以保留 "0-200" 等总区间的数据。
                        if x == '纵向速度区间_total' and '纵向速度区间_total' in class_result.columns:
                            class_result = class_result.rename(columns={'纵向速度区间_total': '纵向速度区间'})
                            # 确保该列在后续处理中为字符串类型
                            class_result['纵向速度区间'] = class_result['纵向速度区间'].astype(str)
                        classification_result.append(class_result)

                if classification_result:
                    all_results.append(pd.concat(classification_result, ignore_index=True))

            if not all_results:
                field_results[tag] = None
                continue

            result = pd.concat(all_results, ignore_index=True)
            
            if result.empty:
                field_results[tag] = None
                continue
                
            result['纵向速度区间'] = result['纵向速度区间'].astype('object')

            result = self._fill_missing_values(result)
            if result is not None and not result.empty:
                result = self._sort_and_format_result(result, classifications, category_order)
                # 给字段添加前缀
                result = self._add_field_prefix(result, tag)
                field_results[tag] = result
            else:
                field_results[tag] = None

        # 合并四个表格
        combined_table = self._combine_error_tables(field_results)
        return combined_table

    def _add_field_prefix(self, df, field_tag):
        """给字段添加前缀"""
        if df.empty:
            return df
        
        # 需要保留的共用列
        common_cols = ['class', 'category', '纵向速度区间']
        
        # 需要添加前缀的列（除了共用列）
        for col in df.columns:
            if col not in common_cols:
                # 提取字段简称（如DX、DVX等）
                field_short = field_tag.split('_')[0]
                df = df.rename(columns={col: f'{field_short}_{col}'})
        
        return df

    def _combine_error_tables(self, field_results):
        """合并四个误差表格"""
        # 提取四个表格（按照固定顺序）
        dx_df = field_results.get('DX_纵向距离误差')
        dvx_df = field_results.get('DVX_纵向速度误差')
        dy_df = field_results.get('DY_横向距离误差')
        dvy_df = field_results.get('DVY_横向速度误差')
        
        # 检查是否有有效的表格
        valid_dfs = []
        for df in [dx_df, dvx_df, dy_df, dvy_df]:
            if df is not None and not df.empty:
                valid_dfs.append(df)
        
        if not valid_dfs:
            return pd.DataFrame()
        
        # 使用第一个表格作为基础
        base_df = valid_dfs[0].copy()
        
        # 合并其他表格
        for i, df in enumerate(valid_dfs[1:], 1):
            # 确定合并键
            merge_keys = ['class', 'category', '纵向速度区间']
            
            # 检查列是否重复
            new_cols = set(df.columns)
            
            # 只合并新列
            cols_to_merge = list(new_cols - set(merge_keys))
            
            if cols_to_merge:
                # 创建只包含合并键和新列的DataFrame
                df_to_merge = df[merge_keys + cols_to_merge].copy()
                
                # 合并到基础表格
                base_df = pd.merge(
                    base_df, 
                    df_to_merge, 
                    on=merge_keys, 
                    how='outer',
                    suffixes=('', f'_{i}')
                )
        
        # 调整列顺序
        return self._reorder_columns(base_df)

    def _fill_missing_values(self, result_combined):
        """填充缺失值"""
        if result_combined is None or result_combined.empty:
            return result_combined
            
        # 确保必要的列存在
        if '纵向速度区间' not in result_combined.columns:
            return result_combined
        
        # 删除纵向速度区间_total列（如果存在）
        if '纵向速度区间_total' in result_combined.columns:
            result_combined = result_combined.drop(columns=['纵向速度区间_total'])
        
        return result_combined

    def _sort_and_format_result(self, result_combined, classifications, category_order):
        """排序和格式化结果"""
        if result_combined is None or result_combined.empty:
            return result_combined
        
        # 强制 category 顺序
        if 'category' in result_combined.columns:
            result_combined["category"] = pd.Categorical(
                result_combined["category"],
                categories=category_order,
                ordered=True
            )
        
        # 定义排序函数，根据class类型获取对应的速度区间顺序
        def custom_sort(row):
            class_val = row['class']
            speed_val = row['纵向速度区间']
            speed_order = SpeedIntervalProcessor.get_speed_order_for_class(class_val)
            
            if speed_val in speed_order:
                return speed_order.index(speed_val)
            else:
                return 999
        
        # 定义class的排序优先级
        class_order = list(classifications.keys())
        
        def class_sort_key(x):
            if x in class_order:
                return class_order.index(x)
            return 999
        
        # 按照 'class', 'category' 和 '纵向速度区间' 进行排序
        result_combined['class_order'] = result_combined['class'].apply(class_sort_key)
        result_combined['sort_order'] = result_combined.apply(custom_sort, axis=1)
        
        # 排序
        result_combined_sorted = result_combined.sort_values(
            by=['class_order', 'category', 'sort_order']
        ).reset_index(drop=True)
        
        if 'category' in result_combined_sorted.columns:
            result_combined_sorted["category"] = result_combined_sorted["category"].astype(str)
        
        # 转换纵向速度区间为字符串（为下一步的有序分类做准备）
        if '纵向速度区间' in result_combined_sorted.columns:
            result_combined_sorted['纵向速度区间'] = result_combined_sorted['纵向速度区间'].astype(str)
        
        # 删除排序列
        columns_to_drop = [col for col in ['sort_order', 'class_order'] if col in result_combined_sorted.columns]
        result_combined_sorted = result_combined_sorted.drop(columns=columns_to_drop)
        
        # 对纵向速度区间进行有序分类设置
        if '纵向速度区间' in result_combined_sorted.columns:
            # 根据class类型确定速度顺序
            speed_order = DataProcessorConfig.VEHICLE_SPEED_ORDER
            
            # 转换为有序分类，'-'会自动被包含
            result_combined_sorted['纵向速度区间'] = pd.Categorical(
                result_combined_sorted['纵向速度区间'],
                categories=speed_order,
                ordered=True
            )
        
        return result_combined_sorted

    def _reorder_columns(self, df):
        """按照定义的顺序重排列"""
        if df.empty:
            return df
        
        # 获取所有现有的列
        existing_cols = list(df.columns)
        
        # 创建一个有序列表，先保留基础列，然后是其他列
        ordered_cols = []
        
        # 先添加基础列（按固定顺序）
        base_cols = ['scene', 'class', 'category', '纵向速度区间']
        for col in base_cols:
            if col in existing_cols:
                ordered_cols.append(col)
                existing_cols.remove(col)
        
        # 然后按照COLUMN_ORDER定义的顺序添加其他列
        column_order = [
            'DX_arc_数量', 'DX_arc_均值', 'DX_arc_误差平均值', 'DX_arc_误差最小值', 'DX_arc_误差最大值', 
            'DX_arc_误差标准差', 'DX_arc_误差占比百分比', 'DX_arc_percentile:0.5', 'DX_arc_percentile:0.9', 'DX_arc_percentile:0.95',
            'DX_KPI需求值', 'DX_KPI达标率',
            'DVX_arc_数量', 'DVX_arc_均值', 'DVX_arc_误差平均值', 'DVX_arc_误差最小值', 'DVX_arc_误差最大值',
            'DVX_arc_误差标准差', 'DVX_arc_误差占比百分比', 'DVX_arc_percentile:0.5', 'DVX_arc_percentile:0.9', 'DVX_arc_percentile:0.95',
            'DVX_KPI需求值', 'DVX_KPI达标率',
            'DY_arc_数量', 'DY_arc_均值', 'DY_arc_误差平均值', 'DY_arc_误差最小值', 'DY_arc_误差最大值',
            'DY_arc_误差标准差', 'DY_arc_误差占比百分比', 'DY_arc_percentile:0.5', 'DY_arc_percentile:0.9', 'DY_arc_percentile:0.95',
            'DY_KPI需求值', 'DY_KPI达标率',
            'DVY_arc_数量', 'DVY_arc_均值', 'DVY_arc_误差平均值', 'DVY_arc_误差最小值', 'DVY_arc_误差最大值',
            'DVY_arc_误差标准差', 'DVY_arc_误差占比百分比', 'DVY_arc_percentile:0.5', 'DVY_arc_percentile:0.9', 'DVY_arc_percentile:0.95',
            'DVY_KPI需求值', 'DVY_KPI达标率'
        ]
        
        for col in column_order:
            if col in existing_cols:
                ordered_cols.append(col)
                existing_cols.remove(col)
        
        # 最后添加任何剩余的列
        ordered_cols.extend(existing_cols)
        
        # 只保留存在的列
        final_cols = [col for col in ordered_cols if col in df.columns]
        
        return df[final_cols]

    def _write_combined_result_to_excel(self, df, writer, sheet_name, is_all=False):
        """写入合并后的Excel结果"""
        if df.empty:
            # 创建空表
            empty_df = pd.DataFrame()
            self._write_result_to_excel(empty_df, writer, sheet_name, is_all)
            return
        
        # 确保最终数据按正确顺序排列
        df = self._ensure_correct_order(df)

        # 添加目标数量列
        if 'class' in df.columns:
            df = self._add_object_num_and_bugscan_columns(df, sheet_name, is_all)

        # 写入数据
        self._write_result_to_excel(df, writer, sheet_name, is_all)
    
    def _add_object_num_and_bugscan_columns(self, df, sheet_name, is_all=False):
        """添加Object_Num和BugScan列"""
        class_idx = df.columns.get_loc('class')
        
        object_nums = []
        bugscan_columns = {col: [] for col in DataProcessorConfig.BUGSCAN_KEYS}
        
        for _, row in df.iterrows():
            class_val = str(row['class'])
            
            # 根据sheet类型确定如何获取数据
            if is_all:
                obj_count, bugscan_sum = self._get_all_data_for_class(class_val)
            elif 'scene' in df.columns and 'case_name' in df.columns:
                case_name = row['case_name']
                scene = row['scene']
                obj_count, bugscan_sum = self._get_scene_data_for_class(case_name, scene, class_val)
            elif 'case_name' in df.columns:
                case_name = row['case_name']
                obj_count, bugscan_sum = self._get_case_data_for_class(case_name, class_val)
            else:
                obj_count, bugscan_sum = 0, {k: 0 for k in DataProcessorConfig.BUGSCAN_KEYS}
            
            object_nums.append(obj_count)
            
            for col_name in DataProcessorConfig.BUGSCAN_KEYS:
                if obj_count == 0:
                    bugscan_columns[col_name].append(bugscan_sum[col_name])
                else:
                    bugscan_columns[col_name].append(f"{round(100 - bugscan_sum[col_name] / obj_count * 100, 2)}%")
        
        # 插入列
        insert_pos = class_idx + 1
        df.insert(insert_pos, 'Object_Num', object_nums)
        insert_pos += 1
        
        for col_name in DataProcessorConfig.BUGSCAN_KEYS:
            df.insert(insert_pos, f"Bug扫描_{col_name}通过率", bugscan_columns[col_name])
            insert_pos += 1
        
        return df
    
    def _get_all_data_for_class(self, class_val):
        """获取所有数据中指定class的数据"""
        obj_count = 0
        bugscan_sum = {k: 0 for k in DataProcessorConfig.BUGSCAN_KEYS}
        
        for case_name, case_data in self.scene_class_target_counts.items():
            for scene, scene_counts in self.scene_class_target_counts[case_name].items():
                if class_val.upper() == "ALL":
                    obj_count += sum(scene_counts.values())
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            obj_count += sum(scene_counts.values())
                        else:
                            sub_class_str = str(sub_class)
                            for key in [sub_class_str, str(sub_class_str)]:
                                if key in scene_counts:
                                    obj_count += scene_counts[key]
                                    break
        
        # BugScan数据
        for case_name, case_data in self.bugscan_target_result.items():
            for scene, bugscan_data in self.bugscan_target_result[case_name].items():
                if class_val.upper() == "ALL":
                    for class_data in self.bugscan_target_result[case_name][scene].values():
                        for k in bugscan_sum:
                            bugscan_sum[k] += class_data.get(k, 0)
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            for class_data in bugscan_data.values():
                                if isinstance(class_data, dict):
                                    for k in bugscan_sum:
                                        bugscan_sum[k] += class_data.get(k, 0)
                        else:
                            sub_class_keys = [str(sub_class), int(sub_class)] if isinstance(sub_class, (int, float)) else [str(sub_class)]
                            found = False
                            for key in sub_class_keys:
                                if key in bugscan_data:
                                    class_data = bugscan_data[key]
                                    if isinstance(class_data, dict):
                                        for k in bugscan_sum:
                                            bugscan_sum[k] += class_data.get(k, 0)
                                    found = True
                                    break
                            
                            if not found:
                                for key in bugscan_data.keys():
                                    if str(sub_class) in str(key):
                                        class_data = bugscan_data[key]
                                        if isinstance(class_data, dict):
                                            for k in bugscan_sum:
                                                bugscan_sum[k] += class_data.get(k, 0)
                                        break
        
        return obj_count, bugscan_sum
    
    def _get_scene_data_for_class(self, case_name, scene, class_val):
        """获取指定case和scene中指定class的数据"""
        obj_count = 0
        bugscan_sum = {k: 0 for k in DataProcessorConfig.BUGSCAN_KEYS}
        
        if case_name in self.scene_class_target_counts:
            if scene in self.scene_class_target_counts[case_name]:
                scene_counts = self.scene_class_target_counts[case_name][scene]
                if class_val.upper() == "ALL":
                    obj_count = sum(scene_counts.values())
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            obj_count += sum(scene_counts.values())
                        else:
                            sub_class_str = str(sub_class)
                            for key in [sub_class_str, str(sub_class_str)]:
                                if key in scene_counts:
                                    obj_count += scene_counts[key]
                                    break
        
        # BugScan数据
        if case_name in self.bugscan_target_result:
            if scene in self.bugscan_target_result[case_name]:
                bugscan_data = self.bugscan_target_result[case_name][scene]
                if class_val.upper() == "ALL":
                    for class_data in self.bugscan_target_result[case_name][scene].values():
                        for k in bugscan_sum:
                            bugscan_sum[k] += class_data.get(k, 0)
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            for class_data in bugscan_data.values():
                                if isinstance(class_data, dict):
                                    for k in bugscan_sum:
                                        bugscan_sum[k] += class_data.get(k, 0)
                        else:
                            sub_class_keys = [str(sub_class), int(sub_class)] if isinstance(sub_class, (int, float)) else [str(sub_class)]
                            found = False
                            for key in sub_class_keys:
                                if key in bugscan_data:
                                    class_data = bugscan_data[key]
                                    if isinstance(class_data, dict):
                                        for k in bugscan_sum:
                                            bugscan_sum[k] += class_data.get(k, 0)
                                    found = True
                                    break
                            
                            if not found:
                                for key in bugscan_data.keys():
                                    if str(sub_class) in str(key):
                                        class_data = bugscan_data[key]
                                        if isinstance(class_data, dict):
                                            for k in bugscan_sum:
                                                bugscan_sum[k] += class_data.get(k, 0)
                                        break
        
        return obj_count, bugscan_sum
    
    def _get_case_data_for_class(self, case_name, class_val):
        """获取指定case中指定class的数据"""
        obj_count = 0
        bugscan_sum = {k: 0 for k in DataProcessorConfig.BUGSCAN_KEYS}
        
        if case_name in self.scene_class_target_counts:
            for scene, scene_counts in self.scene_class_target_counts[case_name].items():
                if class_val.upper() == "ALL":
                    obj_count += sum(scene_counts.values())
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            obj_count += sum(scene_counts.values())
                        else:
                            sub_class_str = str(sub_class)
                            for key in [sub_class_str, str(sub_class_str)]:
                                if key in scene_counts:
                                    obj_count += scene_counts[key]
                                    break
        
        # BugScan数据
        if case_name in self.bugscan_target_result:
            for scene, bugscan_data in self.bugscan_target_result[case_name].items():
                if class_val.upper() == "ALL":
                    for class_data in self.bugscan_target_result[case_name][scene].values():
                        for k in bugscan_sum:
                            bugscan_sum[k] += class_data.get(k, 0)
                else:
                    sub_classes = self.config.get('classifications', {}).get(class_val, [])
                    if not sub_classes:
                        try:
                            sub_classes = [int(class_val)]
                        except ValueError:
                            sub_classes = [class_val]
                    
                    for sub_class in sub_classes:
                        if sub_class == 'ALL':
                            for class_data in bugscan_data.values():
                                if isinstance(class_data, dict):
                                    for k in bugscan_sum:
                                        bugscan_sum[k] += class_data.get(k, 0)
                        else:
                            sub_class_keys = [str(sub_class), int(sub_class)] if isinstance(sub_class, (int, float)) else [str(sub_class)]
                            found = False
                            for key in sub_class_keys:
                                if key in bugscan_data:
                                    class_data = bugscan_data[key]
                                    if isinstance(class_data, dict):
                                        for k in bugscan_sum:
                                            bugscan_sum[k] += class_data.get(k, 0)
                                    found = True
                                    break
                            
                            if not found:
                                for key in bugscan_data.keys():
                                    if str(sub_class) in str(key):
                                        class_data = bugscan_data[key]
                                        if isinstance(class_data, dict):
                                            for k in bugscan_sum:
                                                bugscan_sum[k] += class_data.get(k, 0)
                                        break
        
        return obj_count, bugscan_sum

    def _ensure_correct_order(self, df):
        """确保数据按正确顺序排列（scene → class → category → 纵向速度区间）"""
        if df.empty:
            return df
        
        # 确保纵向速度区间是有序分类类型
        if '纵向速度区间' in df.columns:
            speed_vals = df['纵向速度区间'].unique()
            
            # 使用速度区间顺序
            speed_order = DataProcessorConfig.VEHICLE_SPEED_ORDER
            
            # 确保总区间在最后
            if '0-200' in speed_order:
                speed_order = [s for s in speed_order if s != '0-200'] + ['0-200']
            
            # 添加 '-' 作为缺失值的表示
            if '-' in speed_vals:
                speed_order = ['-'] + speed_order
            
            # 强制转换为有序分类
            df['纵向速度区间'] = pd.Categorical(
                df['纵向速度区间'],
                categories=speed_order,
                ordered=True
            )
        
        # 定义各列的排序顺序
        sort_cols = []
        if 'scene' in df.columns:
            sort_cols.append('scene')
        if 'class' in df.columns:
            sort_cols.append('class')
        if 'category' in df.columns:
            sort_cols.append('category')
        if '纵向速度区间' in df.columns:
            sort_cols.append('纵向速度区间')
        
        # 如果有排序列，则按照这个顺序排序
        if sort_cols:
            df = df.sort_values(by=sort_cols).reset_index(drop=True)
        
        # 转换纵向速度区间回字符串以便输出
        if '纵向速度区间' in df.columns and isinstance(df['纵向速度区间'].dtype, pd.CategoricalDtype):
            df['纵向速度区间'] = df['纵向速度区间'].astype(str)
        
        return df

    def _write_result_to_excel(self, df, writer, sheet_name, is_all=False):
        """写入Excel结果"""
        wb = writer.book
        
        # 处理sheet名称（确保长度不超过31个字符）
        sheet_name = sheet_name[:31]
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
            # 清除所有行
            ws.delete_rows(1, ws.max_row + 1)
        
        # 写入数据前先按case_name分组并排序
        if not df.empty:
            # 创建副本以避免修改原DataFrame
            df_to_write = df.copy()
            
            # 如果不是all结果汇总，需要特殊排序
            if not is_all and 'case_name' in df_to_write.columns:
                df_to_write = self._sort_dataframe(df_to_write)
            
            # 对于需要合并的列，我们要确保分组内的所有值都相同
            df_to_write = self._ensure_group_consistency(df_to_write, is_all)
            
            # 写入数据
            ExcelFormatter.write_dataframe_to_sheet(ws, df_to_write)
        else:
            # 写入空表头
            headers = self._get_sheet_headers(is_all)
            ws.append(headers)
            df_to_write = pd.DataFrame()  # 创建空的df_to_write用于后续处理
        
        # 应用格式
        try:
            # 直接使用 ExcelFormatter 类
            ExcelFormatter.center_align_cells(ws)
            
            # 合并相同单元格
            if not df.empty:
                self._merge_cells_in_sheet(ws, df_to_write, is_all)
        except NameError as e:
            print(f"ExcelFormatter 未找到或导入错误: {e}")
            # 如果没有 ExcelFormatter，可以跳过格式化
            pass
    
    def _sort_dataframe(self, df_to_write):
        """对DataFrame进行排序"""
        # 定义排序函数，提取case_name中的数字部分
        def extract_case_number(case_name):
            """从case_name中提取数字，支持多种格式如test013、case013等"""
            if pd.isna(case_name):
                return 0
            # 提取所有数字
            numbers = re.findall(r'\d+', str(case_name))
            if numbers:
                return int(numbers[0])
            return 0
        
        # 添加临时列用于排序
        df_to_write['_case_num'] = df_to_write['case_name'].apply(extract_case_number)
        
        # 关键修改：先按case_name分组，然后按数字排序
        # 按case_name字符串排序（这会将相同case_name放在一起），然后按数字排序
        sort_cols = ['case_name', '_case_num']
        if 'scene' in df_to_write.columns:
            sort_cols.append('scene')
        sort_cols.append('class')
        
        df_to_write = df_to_write.sort_values(by=sort_cols).reset_index(drop=True)
        
        # 删除临时列
        df_to_write = df_to_write.drop(columns=['_case_num'])
        
        return df_to_write
    
    def _ensure_group_consistency(self, df_to_write, is_all):
        """确保分组内列值一致"""
        bug_cols = [f"Bug扫描_{col}通过率" for col in DataProcessorConfig.BUGSCAN_KEYS]
        
        if is_all:
            group_keys = ['class']
        elif 'scene' in df_to_write.columns:
            group_keys = ['scene', 'class']
        else:
            group_keys = ['case_name', 'class']
        
        for col in bug_cols + (['Object_Num'] if 'Object_Num' in df_to_write.columns else []):
            if col in df_to_write.columns:
                # 获取分组内的第一行的值
                df_to_write[f'{col}_group'] = df_to_write.groupby(group_keys)[col].transform('first')
                df_to_write[col] = df_to_write[f'{col}_group']
                df_to_write = df_to_write.drop(columns=[f'{col}_group'])
        
        return df_to_write
    
    def _get_sheet_headers(self, is_all):
        """获取工作表表头"""
        if is_all:
            return ['class', 'category', '纵向速度区间']
        elif 'scene' in self._df_columns:
            return ['scene', 'class', 'category', '纵向速度区间']
        else:
            return ['case_name', 'class', 'category', '纵向速度区间']
    
    def _merge_cells_in_sheet(self, ws, df_to_write, is_all):
        """合并工作表中的单元格"""
        try:
            # 获取列索引
            col_indices = self._get_column_indices(df_to_write, is_all)
            
            if is_all:
                self._merge_cells_for_all_sheet(ws, df_to_write, col_indices)
            elif 'scene' in df_to_write.columns:
                self._merge_cells_for_scene_sheet(ws, df_to_write, col_indices)
            else:
                self._merge_cells_for_case_sheet(ws, df_to_write, col_indices)
                
        except (KeyError, ValueError, AttributeError) as e:
            # 如果列不存在或格式化器有问题，跳过合并
            print(f"合并单元格时出错: {e}")
            pass
    
    def _get_column_indices(self, df_to_write, is_all):
        """获取列索引"""
        col_indices = {}
        
        if not is_all and 'case_name' in df_to_write.columns:
            col_indices['case_name'] = df_to_write.columns.get_loc('case_name') + 1
        
        if 'scene' in df_to_write.columns:
            col_indices['scene'] = df_to_write.columns.get_loc('scene') + 1
        
        if 'class' in df_to_write.columns:
            col_indices['class'] = df_to_write.columns.get_loc('class') + 1
        
        if 'category' in df_to_write.columns:
            col_indices['category'] = df_to_write.columns.get_loc('category') + 1
        
        # 检查其他列是否存在
        if 'Object_Num' in df_to_write.columns:
            col_indices['Object_Num'] = df_to_write.columns.get_loc('Object_Num') + 1
        
        bug_cols = {}
        for col_name in DataProcessorConfig.BUGSCAN_KEYS:
            col_key = f"Bug扫描_{col_name}通过率"
            if col_key in df_to_write.columns:
                bug_cols[col_key] = df_to_write.columns.get_loc(col_key) + 1
        
        if bug_cols:
            col_indices['bug_cols'] = bug_cols
        
        return col_indices
    
    def _merge_cells_for_all_sheet(self, ws, df_to_write, col_indices):
        """合并all结果汇总工作表的单元格"""
        # 计算分组
        groups = []
        current_group = None
        start_row = 2
        
        for i, (_, row) in enumerate(df_to_write.iterrows(), start=2):
            group_key = (row['class'])
            if group_key != current_group:
                if current_group is not None:
                    groups.append((current_group, start_row, i-1))
                current_group = group_key
                start_row = i
        
        # 添加最后一个分组
        if current_group is not None:
            groups.append((current_group, start_row, i))
        
        # 根据groups合并class列
        for group in groups:
            group_key, start_row, end_row = group
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_indices['class'],
                    end_column=col_indices['class']
                )
        
        # 根据同样的groups合并Object_Num列和bug列
        if 'Object_Num' in col_indices:
            for group in groups:
                group_key, start_row, end_row = group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['Object_Num'],
                        end_column=col_indices['Object_Num']
                    )
        
        # Bug扫描列
        if 'bug_cols' in col_indices:
            for col_idx in col_indices['bug_cols'].values():
                for group in groups:
                    group_key, start_row, end_row = group
                    if start_row < end_row:
                        ws.merge_cells(
                            start_row=start_row,
                            end_row=end_row,
                            start_column=col_idx,
                            end_column=col_idx
                        )
        
        # 合并category列：在每个class分组内部，再按照category合并
        for group in groups:
            group_key, group_start_row, group_end_row = group
            group_df = df_to_write.iloc[group_start_row-2:group_end_row-1]
            # 计算这个分组内的category分组
            cat_groups = []
            current_cat = None
            cat_start_row = group_start_row
            
            for i, (_, row) in enumerate(group_df.iterrows(), start=group_start_row):
                if row['category'] != current_cat:
                    if current_cat is not None:
                        cat_groups.append((current_cat, cat_start_row, i-1))
                    current_cat = row['category']
                    cat_start_row = i
            
            # 添加最后一个category分组
            if current_cat is not None:
                cat_groups.append((current_cat, cat_start_row, i))
            
            # 合并category列
            for cat_group in cat_groups:
                cat_key, start_row, end_row = cat_group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['category'],
                        end_column=col_indices['category']
                    )
    
    def _merge_cells_for_scene_sheet(self, ws, df_to_write, col_indices):
        """合并包含scene的工作表的单元格"""
        # 首先合并 case_name 列（现在相同case_name会在一起）
        # 计算case_name分组
        case_name_groups = []
        current_case_name = None
        start_row = 2
        
        for i, (_, row) in enumerate(df_to_write.iterrows(), start=2):
            if row['case_name'] != current_case_name:
                if current_case_name is not None:
                    case_name_groups.append((current_case_name, start_row, i-1))
                current_case_name = row['case_name']
                start_row = i
        
        # 添加最后一个分组
        if current_case_name is not None:
            case_name_groups.append((current_case_name, start_row, i))
        
        # 合并case_name列
        for case_group in case_name_groups:
            case_key, start_row, end_row = case_group
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_indices['case_name'],
                    end_column=col_indices['case_name']
                )
        
        # 合并scene列（在case_name分组内合并）
        for case_group in case_name_groups:
            case_key, case_start_row, case_end_row = case_group
            case_df = df_to_write.iloc[case_start_row-2:case_end_row-1]
            
            # 计算这个case_name分组内的scene分组
            scene_groups = []
            current_scene = None
            scene_start_row = case_start_row
            
            for i, (_, row) in enumerate(case_df.iterrows(), start=case_start_row):
                if row['scene'] != current_scene:
                    if current_scene is not None:
                        scene_groups.append((current_scene, scene_start_row, i-1))
                    current_scene = row['scene']
                    scene_start_row = i
            
            # 添加最后一个scene分组
            if current_scene is not None:
                scene_groups.append((current_scene, scene_start_row, i))
            
            # 合并scene列
            for scene_group in scene_groups:
                scene_key, start_row, end_row = scene_group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['scene'],
                        end_column=col_indices['scene']
                    )
        
        # 计算scene-class分组（在scene分组内计算）
        groups = []
        current_group = None
        start_row = 2
        
        for i, (_, row) in enumerate(df_to_write.iterrows(), start=2):
            group_key = (row['scene'], row['class'])
            if group_key != current_group:
                if current_group is not None:
                    groups.append((current_group, start_row, i-1))
                current_group = group_key
                start_row = i
        
        # 添加最后一个分组
        if current_group is not None:
            groups.append((current_group, start_row, i))
        
        # 根据groups合并class列
        for group in groups:
            group_key, start_row, end_row = group
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_indices['class'],
                    end_column=col_indices['class']
                )
        
        # 根据同样的groups合并Object_Num列和bug列
        if 'Object_Num' in col_indices:
            for group in groups:
                group_key, start_row, end_row = group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['Object_Num'],
                        end_column=col_indices['Object_Num']
                    )
        
        # Bug扫描列
        if 'bug_cols' in col_indices:
            for col_idx in col_indices['bug_cols'].values():
                for group in groups:
                    group_key, start_row, end_row = group
                    if start_row < end_row:
                        ws.merge_cells(
                            start_row=start_row,
                            end_row=end_row,
                            start_column=col_idx,
                            end_column=col_idx
                        )
        
        # 合并category列：在每个scene-class分组内部，再按照category合并
        for group in groups:
            group_key, group_start_row, group_end_row = group
            group_df = df_to_write.iloc[group_start_row-2:group_end_row-1]
            # 计算这个分组内的category分组
            cat_groups = []
            current_cat = None
            cat_start_row = group_start_row
            
            for i, (_, row) in enumerate(group_df.iterrows(), start=group_start_row):
                if row['category'] != current_cat:
                    if current_cat is not None:
                        cat_groups.append((current_cat, cat_start_row, i-1))
                    current_cat = row['category']
                    cat_start_row = i
            
            # 添加最后一个category分组
            if current_cat is not None:
                cat_groups.append((current_cat, cat_start_row, i))
            
            # 合并category列
            for cat_group in cat_groups:
                cat_key, start_row, end_row = cat_group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['category'],
                        end_column=col_indices['category']
                    )
    
    def _merge_cells_for_case_sheet(self, ws, df_to_write, col_indices):
        """合并case结果汇总工作表的单元格"""
        # 首先合并 case_name 列（现在相同case_name会在一起）
        # 计算case_name分组
        case_name_groups = []
        current_case_name = None
        start_row = 2
        
        for i, (_, row) in enumerate(df_to_write.iterrows(), start=2):
            if row['case_name'] != current_case_name:
                if current_case_name is not None:
                    case_name_groups.append((current_case_name, start_row, i-1))
                current_case_name = row['case_name']
                start_row = i
        
        # 添加最后一个分组
        if current_case_name is not None:
            case_name_groups.append((current_case_name, start_row, i))
        
        # 合并case_name列
        for case_group in case_name_groups:
            case_key, start_row, end_row = case_group
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_indices['case_name'],
                    end_column=col_indices['case_name']
                )
        
        # 计算case_name-class分组（在case_name分组内计算）
        groups = []
        current_group = None
        start_row = 2
        
        for i, (_, row) in enumerate(df_to_write.iterrows(), start=2):
            group_key = (row['case_name'], row['class'])
            if group_key != current_group:
                if current_group is not None:
                    groups.append((current_group, start_row, i-1))
                current_group = group_key
                start_row = i
        
        # 添加最后一个分组
        if current_group is not None:
            groups.append((current_group, start_row, i))
        
        # 根据groups合并class列
        for group in groups:
            group_key, start_row, end_row = group
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_indices['class'],
                    end_column=col_indices['class']
                )
        
        # 根据同样的groups合并Object_Num列和bug列
        if 'Object_Num' in col_indices:
            for group in groups:
                group_key, start_row, end_row = group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['Object_Num'],
                        end_column=col_indices['Object_Num']
                    )
        
        # Bug扫描列
        if 'bug_cols' in col_indices:
            for col_idx in col_indices['bug_cols'].values():
                for group in groups:
                    group_key, start_row, end_row = group
                    if start_row < end_row:
                        ws.merge_cells(
                            start_row=start_row,
                            end_row=end_row,
                            start_column=col_idx,
                            end_column=col_idx
                        )
        
        # 合并category列：在每个scene-class分组内部，再按照category合并
        for group in groups:
            group_key, group_start_row, group_end_row = group
            group_df = df_to_write.iloc[group_start_row-2:group_end_row-1]
            # 计算这个分组内的category分组
            cat_groups = []
            current_cat = None
            cat_start_row = group_start_row
            
            for i, (_, row) in enumerate(group_df.iterrows(), start=group_start_row):
                if row['category'] != current_cat:
                    if current_cat is not None:
                        cat_groups.append((current_cat, cat_start_row, i-1))
                    current_cat = row['category']
                    cat_start_row = i
            
            # 添加最后一个category分组
            if current_cat is not None:
                cat_groups.append((current_cat, cat_start_row, i))
            
            # 合并category列
            for cat_group in cat_groups:
                cat_key, start_row, end_row = cat_group
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=end_row,
                        start_column=col_indices['category'],
                        end_column=col_indices['category']
                    )


def main():
    """主函数"""
    s_time = datetime.datetime.now()

    input_dir = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\CHZ\Southlake\adas_perception_v3.1_SPC030_2m_80\output\shaoyuqi\20260116_V3.1_2M_3.1.27223.1457"
    # 实际匹配后的结果路径 CSV文件所在目录
    # SpecialProject_labelfile = r"F:\Desktop\python_tools\专项\专项KPI\result\cutout前前车_场景筛选结果.xlsx"
    SpecialProject_labelfile = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\YJ\Test_result\专项\第一批\场景筛选结果(1).xlsx"
    bugscan_file = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\YJ\bugsacn\result\专项\20260117_专项KPI\all_result.txt"

    # input_dir = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\CHZ\Southlake\adas_perception_v3.1_SPC030_2m_80\output\shaoyuqi\20260118_V3.1_2M_3.1.27223.1457"
    # # input_dir = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\CHZ\Southlake\adas_perception_v3.1_SPC030_2m_80\output\shaoyuqi\第二批"
    # # 实际匹配后的结果路径 CSV文件所在目录
    # SpecialProject_labelfile = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\YJ\Test_result\专项\第二批\场景筛选结果专项2.xlsx"
    # # SpecialProject_labelfile = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\YJ\Test_result\专项\第二批\场景筛选结果专项2 - 副本 - 副本.xlsx"
    # bugscan_file = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\YJ\Test_result\专项\第二批\all_result.txt"
    
    # 读取标签信息
    labelinfo = LabelInfoProcessor.get_special_project_labelinfo(SpecialProject_labelfile)

    # 回灌的文件列表
    replay_txtlist = LabelInfoProcessor.find_specified_files(input_dir, file_extensions=".txt")

    # 读取bug扫描的结果
    bugscan_result = BugScanProcessor.get_BugScan_result(bugscan_file)
    
    # 统一输出文件
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    output_file = os.path.join(input_dir, f'V3.1.27223.1457_2M_result_{timestamp}.xlsx')
    
    print("正在处理所有数据...")
    
    # 创建处理器并处理数据
    processor = ARCV3DataProcessor(DataProcessorConfig.UNIFIED_CONFIG, replay_txtlist = replay_txtlist)
    processor.process_data(input_dir, output_file, labelinfo, bugscan_result, key=".csv")
    
    print(f"数据处理完成，输出文件：{output_file}")
    
    e_time = datetime.datetime.now()
    print(f"程序运行时间：{e_time - s_time}")


if __name__ == "__main__":
    main()