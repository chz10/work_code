import datetime
import json
import os
import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DataProcessorConfig:
    """配置参数类"""
    DISTANCE_ORDER = ['0-10', '10-30', '30-60', '60-80', '80-100', '100-120', '120-150', '150-200', '0-200']
    LONG_DIST_BINS = [
        (0, 10, "0-10"), (10, 30, "10-30"), (30, 60, "30-60"),
        (60, 80, "60-80"), (80, 100, "80-100"),
        (100, 120, "100-120"), (120, 150, "120-150"),
        (150, 200, "150-200"), (0, 200, "0-200")
    ]
    
    
class LabelInfoProcessor:
    """标签信息处理类"""
    
    @staticmethod
    def get_special_project_labelinfo(input_path):
        if os.path.splitext(input_path)[1] == ".xlsx":
            return LabelInfoProcessor._process_excel_file(input_path)
        else:
            raise ValueError(f"不支持的文件格式：{input_path}")
        
    @staticmethod
    def _process_excel_file(input_path):
        """处理Excel文件"""
        labelinfo_df = pd.read_excel(input_path, sheet_name="Sheet1", dtype=str)
        result = {}
        pre_scene = ""
        
        for index, row in labelinfo_df.iterrows():
            case_name = str(row.get("case_name", "")).strip()
            cur_scene = str(row.get("scene", "")).strip()
            
            # 大类继承
            if not cur_scene or cur_scene == "nan":
                cur_scene = pre_scene
            else:
                pre_scene = cur_scene

            filename = str(row.get("filename", "")).strip()
            video_dir = str(row.get("video_dir", "")).strip()
            start = row.get("start_frame", "")
            end = row.get("end_frame", "")
            target_id = str(row.get("target_id", "")).strip()
            class_key = row.get("class_key", None)
            class_value = row.get("class_value", None)
            
            # 缺少必要字段，跳过
            if not filename or not start or not end:
                continue
            
            # 转为整数
            try:
                start_frame = int(float(start))
                end_frame = int(float(end))
            except ValueError:
                print("⚠️ 跳过无效行，因为帧编号不是数字:", row)
                continue

            # 聚合逻辑
            if case_name not in result:
                result[case_name] = []

            try:
                target_id = int(target_id)
            except ValueError:
                pass
                
            result[case_name].append({
                "line_num": index + 1,
                "case_name": case_name,
                "scene": cur_scene,
                "filename": filename,
                "video_dir": video_dir,
                "start_frame": start_frame,
                "end_frame": end_frame,
                "target_id_list": [target_id],
                "class_key": class_key,
                "class_value": class_value
            })
        
        return result
    
    @staticmethod
    def find_specified_files(input_path, file_extensions=(".h264", ".h265")):
        search_results = {}
        if os.path.isdir(input_path):
            for root, _, files in os.walk(input_path):
                for file in files:
                    if file.lower().endswith(tuple(file_extensions)):
                        filename = os.path.splitext(file)[0]
                        if filename not in search_results:
                            search_results[filename] = os.path.join(root, file)
        elif os.path.isfile(input_path):
            with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
                for idx, line in enumerate(f, start=1):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        filename = os.path.splitext(os.path.basename(line))[0]
                        search_results[filename] = line
                    except json.JSONDecodeError:
                        print(f"⚠️ 第{idx}行不是合法 JSON: {line}")
        else:
            print("素材路径不存在")
        return search_results


class CSVFileHandler:
    """CSV文件处理类"""
    @staticmethod
    def find_csv_files(directory, key=".csv"):
        csv_files = []
        for file_name in os.listdir(directory):
            if file_name.endswith(key):
                csv_files.append(os.path.join(directory, file_name))
        return csv_files
    
    @staticmethod
    def process_all_csv_data(csv_files):
        """
        汇总目录下所有符合条件的 CSV 文件内容到一个 DataFrame 中
        """
        if not csv_files:
            print(f"未找到符合条件的文件。")
            return pd.DataFrame()

        # 汇总所有 CSV 文件内容
        combined_data = pd.DataFrame()
        for csv_file in csv_files:
            print(f"正在读取文件: {csv_file}")
            try:
                data = pd.read_csv(csv_file)
                combined_data = pd.concat([combined_data, data], ignore_index=True)
            except Exception as e:
                print(f"文件 {csv_file} 读取失败: {e}")

        return combined_data


class DataAggregator:
    """数据聚合处理类"""
    
    @staticmethod
    def safe_nan_func(group, func, default=np.nan):
        group = group[~np.isnan(group)]  # 移除 NaN 值
        if len(group) == 0:  # 如果全是 NaN 或数组为空
            return default
        return func(group)
    
    @staticmethod
    def get_percentile(group, percentile):
        """
        计算百分位数
        """
        # 排除 NaN 值
        group = group[~np.isnan(group)]
        if len(group) == 0:  # 如果全是 NaN
            return np.nan
        sorted_group = np.sort(group)
        index = int(np.ceil(len(sorted_group) * percentile)) - 1
        return sorted_group[index]


class ExcelFormatter:
    """Excel格式处理类"""
    
    @staticmethod
    def merge_same_cells(ws, df, col_name, col_index):
        prev_value = None
        start_index = 2  # Excel从第2行开始写数据

        for row_index, value in enumerate(df[col_name], start=2):
            # 判断是否值发生变化
            if value != prev_value:
                if prev_value is not None:
                    # 如果是不同的值，合并上一个区域
                    ws.merge_cells(
                        start_row=start_index,
                        end_row=row_index - 1,
                        start_column=col_index,
                        end_column=col_index
                    )
                    ws.cell(row=start_index, column=col_index).alignment = Alignment(
                        horizontal='center', vertical='center'
                    )

                # 更新起始位置
                start_index = row_index
                prev_value = value

        # 最后一个区域的合并
        ws.merge_cells(
            start_row=start_index,
            end_row=len(df) + 1,
            start_column=col_index,
            end_column=col_index
        )
        ws.cell(row=start_index, column=col_index).alignment = Alignment(
            horizontal='center', vertical='center'
        )
    
    @staticmethod
    def format_diff_columns(ws, df, diff_cols):
        diff_col_indices = [df.columns.get_loc(col) + 1 for col in diff_cols]
        diff_format = '[Red]"↑"0.000;[Green]"↓"0.000;0.000'

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in diff_col_indices:
                cell = row[col_idx - 1]  # openpyxl 的列是从0开始的
                cell.number_format = diff_format
    
    @staticmethod
    def center_align_cells(ws):
        """设置所有单元格居中且自动换行"""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
    
    @staticmethod
    def write_dataframe_to_sheet(ws, df, startrow=0):
        """将DataFrame写入工作表"""
        for r in dataframe_to_rows(df, index=False, header=(startrow == 0)):
            ws.append(r)


class LaneCollisionChecker:
    """主数据处理类"""
    def __init__(self, config=None, replay_txtlist=None):
        self.distancebin_config = config or DataProcessorConfig.DISTANCE_ORDER
        self.replay_txtlist = replay_txtlist

    def process_data(self, input_dir, output_file, labelinfo, key=".csv"):
        # 查找符合条件的 CSV 文件
        csv_files = CSVFileHandler.find_csv_files(input_dir, key=key)
        # 处理所有CSV文件
        df = CSVFileHandler.process_all_csv_data(csv_files)

        all_filter_df = pd.DataFrame()
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            for case_name, label_attr in labelinfo.items():
                for label_data in label_attr:
                    filename = label_data.get("filename")
                    case_name = label_data.get("case_name")
                    video_dir = label_data.get("video_dir")
                    scene = label_data.get("scene")
                    start_frame = label_data.get("start_frame")
                    end_frame = label_data.get("end_frame")
                    target_id_list = label_data.get("target_id_list", [])
                    class_key = label_data.get("class_key", None)
                    class_value = label_data.get("class_value", None)
                    if end_frame - start_frame <= 10:
                        continue # 过滤过短的场景（小于10帧）

                    df_file = df[df["filename"] == f"{filename}.txt"]  # 先按 filename 粗过滤一层，减少数据量
                    # 条件过滤
                    matched_df = df_file[
                        (df_file["frame_index"] >= start_frame) &
                        (df_file["frame_index"] <= end_frame) &
                        (df_file["replay_tid"].isin(target_id_list))
                    ]

                    # matched_df 匹配专项的结果
                    if not matched_df.empty:
                        # 使用 assign 方法为每行添加新列
                        matched_df = matched_df.assign(
                            case_name=case_name,
                            scene=scene,
                            class_key=class_key,
                            class_value=class_value
                        )
                        all_filter_df = pd.concat([all_filter_df, matched_df], ignore_index=True)
                        tid_list = matched_df['tid'].unique().tolist() # 获取matched_df中tid列的所有值（去除重复）
                        # ============================================================
                        # 如果标注得有class_key和class_value，修改matched_df中的class
                        obj_class = self._process_class_key(class_key, matched_df)
                        if obj_class is None:
                            continue

                        # 计算 压线 的数量
                        touchline_count = (matched_df['collision'] == True).sum()
                        # 计算 不压线 的数量
                        notouchline_count = (matched_df['collision'] == False).sum()

                        # ============================================================
                        if "压线" in scene and "不压线" not in scene:
                            TP_num = touchline_count
                            TN_num = 0
                            FP_num = 0
                            FN_num = notouchline_count
                        else:
                            TP_num = 0
                            TN_num = notouchline_count
                            FP_num = touchline_count
                            FN_num = 0
                        
                        attr = {
                            "case_name": case_name,
                            "场景": scene,
                            "文件名": filename,
                            "视频目录": video_dir,
                            "回灌路径": self.replay_txtlist.get(filename),
                            "开始帧": start_frame,
                            "结束帧": end_frame,
                            "标注目标ID": ','.join(map(str, target_id_list)),
                            "匹配目标ID": ','.join(map(str, tid_list)),
                            "目标类别": class_key,
                            "目标类别值": class_value,
                            "统计总帧数": len(matched_df),
                            "标注压线结果压线_TP": TP_num,
                            "标注压线结果未压线_FN": FN_num,
                            "标注未压线结果未压线_TN": TN_num,
                            "标注未压线结果压线_FP": FP_num,
                            "通过率": f"{(TP_num + TN_num) / (TP_num + FN_num + TN_num + FP_num) * 100:.2f}%"
                        }

                        # 1. 将字典转换为 DataFrame（单行数据）
                        attr_df = pd.DataFrame([attr])  # 注意：将字典放入列表中创建单行DataFrame
                        sheet_name = "单个实例结果"
                        # 2. 获取当前写入位置
                        if sheet_name in writer.sheets:
                            startrow = writer.sheets[sheet_name].max_row
                        else:
                            startrow = 0
                        
                        # 3. 写入DataFrame
                        attr_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            startrow=startrow,
                            header=(startrow == 0),  # 只有当startrow为0时才写表头
                        )
            
            # 统计最终的结果
            self.statisticalResults_to_excel(all_filter_df, writer)

    def _process_class_key(self, class_key, matched_df):
        """处理class_key"""
        if class_key:
            if class_key in ["commercial_vehicle_types", "engineering_vehicle_types"]: # 大车
                matched_df['class'] = "大车"
                return "大车"
            elif class_key in ["small_vehicle_types"]: # 小车
                matched_df['class'] = "小车"
                return "小车"
            elif class_key in ["pedestrian_types", "regulatory_targets"]: # 行人, 法规假人
                matched_df['class'] = "行人"
                return "行人"
            elif class_key in ["two_wheeled_vehicle_types"]: # 骑行人
                matched_df['class'] = "骑行人"
                return "骑行人"
            elif class_key in ["special_vehicle_types"]: # 特种车
                matched_df['class'] = "特种车"
                return "特种车"
            else:
                print("不支持的class_key:", class_key)
                return None
        else:
            # 使用算法输出的class作为目标类型计算
            return None
    
    
    def statisticalResults_to_excel(self, all_filter_df, writer):
        case_namelist = all_filter_df['case_name'].unique().tolist()
        scene_list = all_filter_df['scene'].unique().tolist()
        class_key_list = all_filter_df['class_key'].unique().tolist()
        longdis_bin_config = DataProcessorConfig.LONG_DIST_BINS

        # ============================================================
        # 细分到case_name, scene, class_key
        all_states_list = []
        for case_name in case_namelist:
            case_df = all_filter_df[all_filter_df['case_name'] == case_name]
            for scene in scene_list:
                scene_df = case_df[case_df['scene'] == scene]
                for class_key in class_key_list:
                    class_df = scene_df[scene_df['class_key'] == class_key]
                    for dist_bin in longdis_bin_config:
                        min_dist = dist_bin[0]
                        max_dist = dist_bin[1]
                        dist_bin_name = dist_bin[2]
                        dist_bin_df = class_df[(case_df['x'] >= min_dist) & (case_df['x'] <= max_dist)]
                        # ============================================================
                        # 计算 压线 的数量
                        touchline_count = (dist_bin_df['collision'] == True).sum()
                        # 计算 不压线 的数量
                        notouchline_count = (dist_bin_df['collision'] == False).sum()
                        # ============================================================
                        if "压线" in scene and "不压线" not in scene:
                            TP_num = touchline_count
                            TN_num = 0
                            FP_num = 0
                            FN_num = notouchline_count
                        else:
                            TP_num = 0
                            TN_num = notouchline_count
                            FP_num = touchline_count
                            FN_num = 0
                        if TP_num + FN_num + TN_num + FP_num == 0:
                            continue
                        attr = {
                            "case_name": case_name,
                            "场景": scene,
                            "目标类别": class_key,
                            "距离区间": dist_bin_name,
                            "统计总帧数": len(dist_bin_df),
                            "标注压线结果压线_TP": TP_num,
                            "标注压线结果未压线_FN": FN_num,
                            "标注未压线结果未压线_TN": TN_num,
                            "标注未压线结果压线_FP": FP_num,
                            "通过率": f"{(TP_num + TN_num) / (TP_num + FN_num + TN_num + FP_num) * 100:.2f}%"
                        }
                        all_states_list.append(attr)

        # 1. 将字典转换为 DataFrame（单行数据）
        attr_df = pd.DataFrame(all_states_list)  # 注意：将字典放入列表中创建单行DataFrame
        sheet_name = "统计结果"
        # 2. 获取当前写入位置
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0
        
        # 3. 写入DataFrame
        attr_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=startrow,
            header=(startrow == 0),  # 只有当startrow为0时才写表头
        )

        # ============================================================
        # 细分到case_name, class_key
        all_states_list = []
        for case_name in case_namelist:
            case_df = all_filter_df[all_filter_df['case_name'] == case_name]
            for class_key in class_key_list:
                class_df = case_df[case_df['class_key'] == class_key]
                for dist_bin in longdis_bin_config:
                    min_dist = dist_bin[0]
                    max_dist = dist_bin[1]
                    dist_bin_name = dist_bin[2]
                    dist_bin_df = class_df[(class_df['x'] >= min_dist) & (class_df['x'] <= max_dist)]
                    TP_num = 0
                    TN_num = 0
                    FP_num = 0
                    FN_num = 0
                    total_num = 0
                    for scene in scene_list:
                        scene_df = dist_bin_df[dist_bin_df['scene'] == scene]
                        total_num += len(scene_df)
                        # ============================================================
                        # 计算 压线 的数量
                        touchline_count = (scene_df['collision'] == True).sum()
                        # 计算 不压线 的数量
                        notouchline_count = (scene_df['collision'] == False).sum()
                        # ============================================================
                        if "压线" in scene and "不压线" not in scene:
                            TP_num += touchline_count
                            TN_num += 0
                            FP_num += 0
                            FN_num += notouchline_count
                        else:
                            TP_num += 0
                            TN_num += notouchline_count
                            FP_num += touchline_count
                            FN_num += 0
                    if TP_num + FN_num + TN_num + FP_num == 0:
                        continue
                    attr = {
                        "case_name": case_name,
                        "目标类别": class_key,
                        "距离区间": dist_bin_name,
                        "统计总帧数": total_num,
                        "标注压线结果压线_TP": TP_num,
                        "标注压线结果未压线_FN": FN_num,
                        "标注未压线结果未压线_TN": TN_num,
                        "标注未压线结果压线_FP": FP_num,
                        "通过率": f"{(TP_num + TN_num) / (TP_num + FN_num + TN_num + FP_num) * 100:.2f}%"
                    }
                    all_states_list.append(attr)

        # 1. 将字典转换为 DataFrame（单行数据）
        attr_df = pd.DataFrame(all_states_list)  # 注意：将字典放入列表中创建单行DataFrame
        sheet_name = "统计结果_case_name"
        # 2. 获取当前写入位置
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0
        
        # 3. 写入DataFrame
        attr_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=startrow,
            header=(startrow == 0),  # 只有当startrow为0时才写表头
        )

        # ============================================================
        # class_key
        all_states_list = []
        for dist_bin in longdis_bin_config:
            min_dist = dist_bin[0]
            max_dist = dist_bin[1]
            dist_bin_name = dist_bin[2]
            dist_bin_df = all_filter_df[(all_filter_df['x'] >= min_dist) & (all_filter_df['x'] <= max_dist)]
            TP_num = 0
            TN_num = 0
            FP_num = 0
            FN_num = 0
            total_num = 0
            for class_key in class_key_list:
                class_df = dist_bin_df[dist_bin_df['class_key'] == class_key]
                for scene in scene_list:
                    scene_df = class_df[class_df['scene'] == scene]
                    total_num += len(scene_df)
                    # ============================================================
                    # 计算 压线 的数量
                    touchline_count = (scene_df['collision'] == True).sum()
                    # 计算 不压线 的数量
                    notouchline_count = (scene_df['collision'] == False).sum()
                    # ============================================================
                    if "压线" in scene and "不压线" not in scene:
                        TP_num += touchline_count
                        TN_num += 0
                        FP_num += 0
                        FN_num += notouchline_count
                    else:
                        TP_num += 0
                        TN_num += notouchline_count
                        FP_num += touchline_count
                        FN_num += 0
            if TP_num + FN_num + TN_num + FP_num == 0:
                continue
            attr = {
                "目标类别": "ALL",
                "距离区间": dist_bin_name,
                "统计总帧数": total_num,
                "标注压线结果压线_TP": TP_num,
                "标注压线结果未压线_FN": FN_num,
                "标注未压线结果未压线_TN": TN_num,
                "标注未压线结果压线_FP": FP_num,
                "通过率": f"{(TP_num + TN_num) / (TP_num + FN_num + TN_num + FP_num) * 100:.2f}%"
            }
            all_states_list.append(attr)

        for class_key in class_key_list:
            class_df = all_filter_df[all_filter_df['class_key'] == class_key]
            for dist_bin in longdis_bin_config:
                min_dist = dist_bin[0]
                max_dist = dist_bin[1]
                dist_bin_name = dist_bin[2]
                dist_bin_df = class_df[(class_df['x'] >= min_dist) & (class_df['x'] <= max_dist)]
                TP_num = 0
                TN_num = 0
                FP_num = 0
                FN_num = 0
                total_num = 0
                for scene in scene_list:
                    scene_df = dist_bin_df[dist_bin_df['scene'] == scene]
                    total_num += len(scene_df)
                    # ============================================================
                    # 计算 压线 的数量
                    touchline_count = (scene_df['collision'] == True).sum()
                    # 计算 不压线 的数量
                    notouchline_count = (scene_df['collision'] == False).sum()
                    # ============================================================
                    if "压线" in scene and "不压线" not in scene:
                        TP_num += touchline_count
                        TN_num += 0
                        FP_num += 0
                        FN_num += notouchline_count
                    else:
                        TP_num += 0
                        TN_num += notouchline_count
                        FP_num += touchline_count
                        FN_num += 0
                if TP_num + FN_num + TN_num + FP_num == 0:
                    continue
                attr = {
                    "目标类别": class_key,
                    "距离区间": dist_bin_name,
                    "统计总帧数": total_num,
                    "标注压线结果压线_TP": TP_num,
                    "标注压线结果未压线_FN": FN_num,
                    "标注未压线结果未压线_TN": TN_num,
                    "标注未压线结果压线_FP": FP_num,
                    "通过率": f"{(TP_num + TN_num) / (TP_num + FN_num + TN_num + FP_num) * 100:.2f}%"
                }
                all_states_list.append(attr)

        # 1. 将字典转换为 DataFrame（单行数据）
        attr_df = pd.DataFrame(all_states_list)  # 注意：将字典放入列表中创建单行DataFrame
        sheet_name = "统计结果class"
        # 2. 获取当前写入位置
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0
        
        # 3. 写入DataFrame
        attr_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=startrow,
            header=(startrow == 0),  # 只有当startrow为0时才写表头
        )


def main():
    """主函数"""
    s_time = datetime.datetime.now()

    input_dir = r"\\hz-iotfs02\Model_Test\TestSpace\Personal_Space\CHZ\Southlake\adas_perception_v3.1_SPC030_2m_80\output\shaoyuqi\20260116_V3.1_2M_3.1.27223.1457"
    # 结果匹配后的结果路径 CSV文件所在目录
    SpecialProject_labelfile = r"F:\Desktop\python_tools\专项\目标与车道线横向位置关系专项\result\目标与车道线关系_场景筛选结果.xlsx"
    # SpecialProject_labelfile = r"F:\Desktop\python_tools\专项\目标与车道线横向位置关系专项\result\目标与车道线关系_场景筛选结果20260127 - 副本.xlsx"

    # 读取标签信息
    labelinfo = LabelInfoProcessor.get_special_project_labelinfo(SpecialProject_labelfile)
    # 回灌的文件列表
    replay_txtlist = LabelInfoProcessor.find_specified_files(input_dir, file_extensions=".txt")

    # 统一输出文件
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    output_file = os.path.join(input_dir, f'目标与车道线关系_result_{timestamp}.xlsx')
    
    print("正在处理所有数据...")
    
    # 创建处理器并处理数据
    processor = LaneCollisionChecker(DataProcessorConfig.DISTANCE_ORDER, replay_txtlist = replay_txtlist)
    processor.process_data(input_dir, output_file, labelinfo, key="vision_replay_match_part1.csv")
    
    print(f"数据处理完成，输出文件：{output_file}")
    
    e_time = datetime.datetime.now()
    print(f"程序运行时间：{e_time - s_time}")


if __name__ == "__main__":
    main()